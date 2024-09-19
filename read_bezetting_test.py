# GENERATE BUSINESS INTELLIGENCE COMING FROM A PLANNING SHEET

from datetime import datetime
import string
import pandas as pd

from openpyxl import load_workbook
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import urllib.request

test = False  # To test or not to test (to see if the fillcolors in the sheet are right.)

local = True
if local:
    #excel_file = r"C:\Users\rcxsm\Documents\python_scripts\streamlit_scripts\input\dummy_occupation.xlsx"
    excel_file = r"https://github.com/rcsmit/streamlit_scripts/blob/main/input/dummy_occupation.xlsx?raw=true"
    
    wb = load_workbook(excel_file, data_only=True)
else:
    excel_file = r"https://github.com/rcsmit/streamlit_scripts/blob/main/input/dummy_occupation.xlsx?raw=true"
    urllib.request.urlretrieve(excel_file, "test.xlsx")
    wb = load_workbook("test.xlsx", data_only=True)


start_month, end_month = 3, 10
year_ = [2022]
(start_month,end_month) = st.sidebar.slider("Months (from/until (incl.))", 1, 12, (3,10))

if start_month > end_month:
    st.warning("Make sure that the end month is not before the start month")
    st.stop()

year = st.sidebar.multiselect("Years", year_, year_)
selection_list_accos_ = ['ALPHA' ,'BRAVO']
selection_list_accos = st.sidebar.multiselect("Which acco's", selection_list_accos_,selection_list_accos_)
sh_2022 = wb["TEST"]


    # rij met de name, startrow, eindrow
to_do_2022 = [
    [1, 4, 9],  # bal
    [11, 14, 23],  # wai

] 


def find_fill_color(cell):
    """Find fill color of a cell.
    Hulproutine, wordt niet aangeroepen in het script

    # dirty solution to find the fill color.
    # as solutions given here https://stackoverflow.com/questions/58429823/getting-excel-cell-background-themed-color-as-hex-with-openpyxl
    # dont work

    Args:
        cell (string): The cell you want to find the color from
    """

    val = sh_2022[cell].fill.start_color.rgb

    try:
        valx = val[0]
        valx = val
    except:
        valx = sh_2022[cell].fill.start_color.theme

    theme = sh_2022[cell].fill.start_color.theme
    tint = sh_2022[cell].fill.start_color.tint
    st.write(f"Valx = {valx  } | Theme = {theme} ||Tint = {tint}")

    # DOESNT WORK
    # val = int(sh_2022[cell].fill.start_color.index, 16)
    # st.write (val)
    # #hex_color = "%06x" % (val && 0xFFFFFF)
    # st.write(hex_color)
    

def retrieve_prijzen():
    """Retrieve the average price for an accomodation in a given month
    Returns:
        df: Table with the prices
    """    
    sheet_id_prijzen = "1eNdn6mAglADaqOuRyQAHkx-yFgd-2mlK_fACLIHoJNk"
    sheet_name_prijzen = "TEST"
    url_prijzen = f"https://docs.google.com/spreadsheets/d/{sheet_id_prijzen}/gviz/tq?tqx=out:csv&sheet={sheet_name_prijzen}"

    df_prijzen = pd.read_csv(url_prijzen, delimiter=",")
    # df_prijzen_stacked = df_prijzen.stack()
    df_prijzen_stacked = df_prijzen.melt(
        "acco_type", var_name="maand_int", value_name="price_per_night"
    )
    df_prijzen_stacked["maand_str"] = df_prijzen_stacked["maand_int"].astype(str)
    # .set_index('acco_type').stack().rename(columns={'price_per_night':'month'})
    return df_prijzen_stacked

def create_check_table_per_accotype(df):
    """Generate tables per accotype to see if the sheet is 100% correct (fill colors right*).
       The last column has to be zero

        * a booking starts with green
        * a booking end with cyaan (wissel) or red (checkout)
    Args:
        df (_type_): _description_
    """
    list_of_accotypes_ = df.acco_type.unique()
    list_of_accotypes = [list_of_accotypes_[4]]  # if you only one acco type
    for acco in list_of_accotypes:
        df_acco = df[df["acco_type"] == acco].reset_index()

        df_acco = df_acco.assign(bezet_saldo=None)
        df_acco.loc[0, "bezet_saldo"] = 0
        df_acco["bezet_theorie"] = (
            df_acco["geel"] + df_acco["wissel"] + df_acco["new_arrival"]
        )
        for i in range(1, len(df_acco)):
            df_acco.loc[i, "bezet_saldo"] = (
                df_acco.loc[i - 1, "bezet_saldo"]
                + df_acco.loc[i, "new_arrival"]
                - df_acco.loc[i, "vertrek_no_clean"]
                - df_acco.loc[i, "vertrek_clean"]
            )
        df_acco["verschil_bezet"] = df_acco["bezet_theorie"] - df_acco["bezet_saldo"]
        st.write(df_acco.to_string())
        st.write(df_acco.sum())


def create_table_per_accotype_per_month(df, year):
    # list_of_accotypes = df.acco_type.unique()
    # # list_of_accotypes = [list_of_accotypes_[8]] #if you only one acco type

    mm = ["month"]

    columns = [*mm, *selection_list_accos]
    what = ["omzet", "verblijfsduur", "bezetting","aantal_overnachtingen","aantal_boekingen"]
    months = [
        "januari",
        "februari",
        "maart",
        "april",
        "mei",
        "juni",
        "juli",
        "augustus",
        "september",
        "oktober",
        "november",
        "december",
    ]
    for w in what:
        list_ = []
        for m in range(start_month, end_month + 1):
            rij = [months[m - 1]]
            for acco in selection_list_accos:
                
                df_acco = df[
                    (df["acco_type"] == acco) & (df["maand_str"] == str(m))
                ].reset_index()
                
                in_tabel_te_plaatsen = round(print_businessinfo(df_acco, w, year), 1)
                rij.append(in_tabel_te_plaatsen)

            list_.append(rij)

        df__ = pd.DataFrame(list_, columns=columns)
        st.write()
        st.write(f"{w} -- {year} ---")
        if w == "omzet":
            df__["Total"] = df__.sum(numeric_only=True, axis=1)
            row_sum = df__.iloc[:, 1:].sum()
            df__.loc["Total"] = row_sum
        st.write(df__.transpose().astype(str))


def print_businessinfo(df_acco, what, year):
    """print and return the business intelligence
    Args:
        df_acco (_type_): _description_
        what (_type_): _description_

    Returns:
        _type_: _description_
    """
   
    aantal_boekingen =  df_acco["wissel"].sum() + df_acco["new_arrival"].sum()
    if what == None:
        st.write(f"-----{year}-----")
        
        st.write(
            f"Aantal boekingen {aantal_boekingen}"
        )
        
        # st.write (f'Aantal accos {df_acco["aantal"].mean()} | aantal nachten {len(df_acco)}')
    if (df_acco["aantal"].mean() * len(df_acco) - df_acco["out_of_order"].sum()) != 0:
        bezetting = round(
            (
                df_acco["geel"].sum()
                + df_acco["wissel"].sum()
                + df_acco["new_arrival"].sum()
            )
            / (
                (df_acco["aantal"].mean() * len(df_acco))
                - df_acco["out_of_order"].sum()
            )
            * 100,
            2,
        )
        if what == None:
            st.write(f"Bezetting {bezetting} %")
    else:
        bezetting = 0
    aantal_overnachtingen = df_acco["geel"].sum()  + df_acco["wissel"].sum()   + df_acco["new_arrival"].sum()
            
    if (df_acco["wissel"].sum() + df_acco["new_arrival"].sum()) != 0:
        # De verblijfsduur is vertekend als je het per maand, per acco bekijkt in rustige maanden, zie bijv. bali, september 2019 (maar 1 aankomst, maar mensen die nog vanuit augustus aanwezig zijn)
        verblijfsduur = round(
            (
                df_acco["geel"].sum()
                + df_acco["wissel"].sum()
                + df_acco["new_arrival"].sum()
            )
            / (df_acco["wissel"].sum() + df_acco["new_arrival"].sum()),
            2,
        )
        if what == None:
            st.write(f"Gemiddeld verblijfsduur {verblijfsduur}")
    else:
        verblijfsduur = 0
    omzet = df_acco["omzet"].sum()
    aantal_acco = 0
    if year == 2019:
        to_do = to_do_2019
        sh = sh_2019
    if year == 2021:
        to_do = to_do_2021
        sh = sh_2021
    if year == 2022:
        to_do = to_do_2022
        sh = sh_2022
    for t in to_do:
        a_t = str(sh["a" + str(t[0])].value)
        if a_t in selection_list_accos:
            aantal_acco +=(t[2]-t[1]+1)
    if what == None:
        st.write(f"Totale omzet: {omzet}")
    if what == "omzet":
        to_return = omzet
    elif what == "verblijfsduur":
        to_return = verblijfsduur
    elif what == "bezetting":
    
        to_return = bezetting
    elif what == "aantal_boekingen":
        to_return  = aantal_boekingen
    elif what =='aantal_overnachtingen':
        to_return  = aantal_overnachtingen
    else:
        to_return = [year, omzet,aantal_acco, round((omzet/aantal_acco),2),round(bezetting,1),aantal_boekingen,round(verblijfsduur,1),aantal_overnachtingen,round(aantal_overnachtingen/aantal_acco,1)]
    return to_return

def generate_columns_to_use():
    """Generate a list with columns to use, eg. from A to ZZ

    Returns:
        _type_: _description_
    """
    alphabet = list(string.ascii_uppercase)
    alphabet_tot_h = [
        "",
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
    ]
    alphabet_to_use = []
    for alp in alphabet_tot_h:
        for alp2 in alphabet:
            alphabet_to_use.append(alp + alp2)
    alphabet_to_use = alphabet_to_use[1:]
    return alphabet_to_use


def add_row(list, acco_type,acco_number, guest_name, checkin_date, checkout_date):
    """Add a row to the list.

    Args:
        list (list): the list where the new row has to be added
        acco_type (str): _description_
        acco_number (str): _description_
        guest_name (str): _description_
        checkin_date (str): _description_
        checkout_date (str): _description_

    Returns:
        list: _description_
    """    
    number_of_days = datetime.strptime(checkout_date, "%Y-%m-%d") - datetime.strptime(checkin_date, "%Y-%m-%d") 
    list.append([acco_type,acco_number, guest_name, checkin_date, checkout_date, number_of_days])
    return list

def make_booking_table(year):
    """Generate a booking_tabel from the booking chart
    columns: [acco_type,acco_number, guest_name, checkin_date, checkout_date, number_of_days]

    Args:
        year (int): the year of which you want the booking table
    """
    year__ = year
    columns_to_use = generate_columns_to_use()
  
    st.subheader("Booking table")
    if year == 2022:
        to_do = to_do_2022
        sh = sh_2022
    list =[]
  
    for t in to_do:
        
        acco_type = str(sh["a" + str(t[0])].value)
        for r in range(t[1],t[2]+1):
            print (r)
            acco_number_cell = "a" + str(r)
            acco_number = str(sh[acco_number_cell].value)
            for c in columns_to_use:
                
                cell_ = c +str(r)
                datum_ = str(sh[c + "2"].value)
                try:
                    datum2 = datetime.strptime(datum_, "%Y-%m-%d %M:%H:%S")
                    datum = datetime.strftime(datum2, "%Y-%m-%d")
                    month = datum2.month
                    year = datum2.year
                except:
                    datum = None
                    month = None
                    year = None

                val = sh[cell_].fill.start_color.rgb
                try:
                    valx = val[0]
                    valx = val
                except:
                    valx = sh[cell_].fill.start_color.theme
                
                if valx == 9:  # licht groen
                    checkin_date  = datum
                    guest_name = str(sh[c + str(r)].value)
                   
                    
                elif valx == "FFFF0000":  # rood
                    checkout_date  = datum
                    
                    list = add_row(list,acco_type,acco_number, guest_name, checkin_date, checkout_date)
                elif valx == "FF7030A0":  # paars
                    checkout_date  = datum
                    list = add_row(list, acco_type,acco_number, guest_name, checkin_date, checkout_date)
                elif valx == 5:  # bruin
                    pass
                elif valx == 0 or valx == 6:  # zwart of grijs
                    pass
                elif valx == "FF00B0F0":  # lichtblauw / cyaan
                    checkout_date  = datum
                    list = add_row(list,acco_type,acco_number, guest_name, checkin_date, checkout_date)
                    checkin_date  = datum
                    guest_name = str(sh[c + str(r)].value)

                elif valx == "FFFFFF00":  # geel / bezet
                    pass

                
    df = pd.DataFrame(
        list,
        columns=[
            "acco_type","acco_number", 
            "guest_name",
            "checkin_date",
            "checkout_date","number_of_days"
        ],
    )
    df['number_of_days'] = df['number_of_days'].dt.days.astype('int16')
    st.subheader(f"Distribution of length of stay in {year__}")
    st.write (df)
    st.write (f"Number of days total : {df['number_of_days'].sum()}")
    st.write (f"Number of days min : {df['number_of_days'].min()}")
    st.write (f"Number of days max : {df['number_of_days'].max()}")
    st.write (f"Number of days average : {df['number_of_days'].mean()}")


    freq_tabel = df['number_of_days'].value_counts()
    fig = px.histogram(df, x="number_of_days")
    #plotly.offline.plot(fig)
    st.plotly_chart(fig, use_container_width=True)
    st.write("Frequency table")
    st.write (freq_tabel)
  

def make_complete_df(columns_to_use, year):
    """Generate the dataframe
        Columns: ['acco_type', 'aantal', 'datum',"month","year", "new_arrival","vertrek_no_clean", "vertrek_clean", "wissel", "geel"])

    Args:
        columns_to_use (list with strings): which columns to scrape, eg. from "A to ... ZZ "

    Returns:
        df: dataframe
    """
    list_complete = []
    if year == 2022:
        to_do = to_do_2022
        sh = sh_2022

    for a in columns_to_use:
        #       [rij met de naam, start, eind]

        for t in to_do:
            acco_type = str(sh["a" + str(t[0])].value)
            ii = []
            for x in range(t[1], t[2] + 1):
                ii.append(a + str(x))
            bezet = 0
            aantal = t[2] - t[1] + 1
            vertrek_no_clean = 0
            vertrek_clean = 0
            vertrek_totaal = 0
            wissel = 0
            geel = 0
            out_of_order = 0
            new_arrival = 0
            try:
                datum = str(sh[a + "2"].value)

                datum2 = datetime.strptime(datum, "%Y-%m-%d %M:%H:%S")
                datum3 = datetime.strftime(datum2, "%Y-%m-%d")
                month = datum2.month
                year = datum2.year
            except:
                datum3 = "STOP"

            if datum3 != "STOP":
                for i in ii:
                    val = sh[i].fill.start_color.rgb
                    try:
                        valx = val[0]
                        valx = val
                    except:
                        valx = sh[i].fill.start_color.theme

                    if valx == "FFFF0000":  # rood
                        vertrek_no_clean += 1
                    elif valx == "FF7030A0":  # paars
                        vertrek_clean += 1
                    elif valx == 5:  # bruin
                        vertrek_totaal += 1
                    elif valx == 0 or valx == 6:  # zwart of grijs
                        out_of_order += 1
                    elif valx == "FF00B0F0":  # lichtblauw / cyaan
                        wissel += 1
                    elif valx == "FFFFFF00":  # geel / bezet
                        geel += 1
                    elif valx == 9:  # licht groen
                        new_arrival += 1

                row = [
                    acco_type,
                    aantal,
                    datum3,
                    month,
                    year,
                    new_arrival,
                    vertrek_no_clean,
                    vertrek_clean,
                    wissel,
                    geel,
                    out_of_order,
                ]
                list_complete.append(row)
    df = pd.DataFrame(
        list_complete,
        columns=[
            "acco_type",
            "aantal",
            "datum",
            "month",
            "year",
            "new_arrival",
            "vertrek_no_clean",
            "vertrek_clean",
            "wissel",
            "geel",
            "out_of_order",
        ],
    )
    df["in_house"] = df["geel"] + df["new_arrival"] + df["wissel"]
    df["maand_str"] = df["month"].astype(str)
    df = df[(df["month"] >= start_month) & (df["month"] <= end_month)]
    
    
    df = df[df["acco_type"].isin(selection_list_accos)]
    st.write(df)
    return df

def make_date_columns(df):
    df['datum'] = pd.to_datetime(df.datum, format='%Y-%m-%d')
    df["jaar"] = df["datum"].dt.strftime("%Y")
    df["maand"] = df["datum"].dt.strftime("%m").astype(str).str.zfill(2)
    df["dag"] = df["datum"].dt.strftime("%d").astype(str).str.zfill(2)
    df["maand_dag"] = df["maand"] + "-" + df["dag"]
    df["dag_maand"] = df["dag"] + "-" + df["maand"]
    df["datum_str"] = df["datum"].astype(str)
    df["datum_"] = pd.to_datetime(df["maand_dag"], format="%m-%d")
    return df


def  make_occopuation_graph_per_acco(df_):
    
    for y in year:
        df_all_years_grouped = df_[df_["jaar"] == str(y)]
        
        data = []
        fig = go.Figure()
        df_all_years_grouped = df_all_years_grouped.sort_values(by='datum') 
        df_all_years_pivot_a = df_all_years_grouped.pivot(index=['maand_dag','datum_'], columns='acco_type', values='occupation').reset_index()
        
        width, opacity = 1,1
        for a in selection_list_accos:
            try:
                # niet alle accomodaties zijn gebruikt in alle jaren

                points = go.Scatter(
                                name = a,
                                x=df_all_years_pivot_a["datum_"],
                                y=df_all_years_pivot_a[a],
                                line=dict(width=width),
                                opacity=opacity,
                                mode="lines",
                            )
                    
                data.append(points)
            except:
                pass
        layout = go.Layout(
                    yaxis=dict(title=f"Bezetting (%)"),
                    title=f"Bezetting per acco type ({a}) in {y}",
                )
        fig = go.Figure(data=data, layout=layout)
        fig.update_layout(xaxis=dict(tickformat="%d-%m"))
        #fig.show()
        #plotly.offline.plot(fig)
        st.plotly_chart(fig, use_container_width=True)
        #st.write(df_all_years_pivot)

def  make_occopuation_graph(df_all_years_grouped):
    data = []
    fig = go.Figure()
    df_all_years_grouped = df_all_years_grouped.sort_values(by='datum') 
    df_all_years_grouped["jaar"] = df_all_years_grouped["jaar"].astype(str)
    df_all_years_pivot = df_all_years_grouped.pivot(index=['maand_dag','datum_'], columns='jaar', values='occupation').reset_index()
    #st.write(df_all_years_pivot) 
    for y in year:
        y_=str(y)
        if y_ == "2022":
            width = 2
            opacity = 1
        else:
            width = 0.7
            opacity = 0.8

        points = go.Scatter(
                        name = y_,
                        x=df_all_years_pivot["datum_"],
                        y=df_all_years_pivot[y_],
                        line=dict(width=width),
                        opacity=opacity,
                        mode="lines",
                    )
            
        data.append(points)
    layout = go.Layout(
                yaxis=dict(title=f"Bezetting (%)"),
                title="Bezetting",
            )
    fig = go.Figure(data=data, layout=layout)
    fig.update_layout(xaxis=dict(tickformat="%d-%m"))
    #fig.show()
    #plotly.offline.plot(fig)
    st.plotly_chart(fig, use_container_width=True)
    #st.write(df_all_years_pivot)
    
def most_checkins_out(df_all_years_grouped):

    for y in year:
        df_all_years_grouped_one_year = df_all_years_grouped[df_all_years_grouped["jaar"]==str(y)]
        df_checkouts = df_all_years_grouped_one_year.sort_values(by='checkouts',ascending=False)
        checkouts = df_checkouts[["datum", "checkouts"]]
        df_checkins = df_all_years_grouped_one_year.sort_values(by='checkins', ascending=False) 
        checkins = df_checkins[["datum", "checkins"]]
        st.subheader (y)
        col1,col2=st.columns(2)
        with col1:
            st.subheader("Checkouts")
            st.write (checkouts.head(10))
            fig = px.histogram(df_checkouts, x="checkouts", title=y)
            st.plotly_chart(fig, use_container_width=True)
        with col2:
            st.subheader("Checkins")
            st.write (checkins.head(10))
            fig = px.histogram(df_checkins, x="checkins", title=y)
            st.plotly_chart(fig, use_container_width=True)

def calculate_occupation(df):
    
    df["checkins"] = df["wissel"] + df["new_arrival"]
    df["checkouts"] = df["wissel"] + df["vertrek_no_clean"]+ df["vertrek_clean"]
    df["occupation"] = round(100*(df["new_arrival"]+df["wissel"]+df["geel"]) /
                                                (df["aantal"]-df["out_of_order"]),2)
    df= make_date_columns(df)
    return df


def generate_and_show_info_all_years():
    df_all_years =  pd.DataFrame()

    list,columns = [], ["year","omzet_eur","aantal_acco","omzet_per_acco","bezetting_%","aantal_boekingen","gem_verblijfsduur","aantal_overnachtingen","aantal_overnachtingen_per_acco"]
    for y in year:
        columns_to_use = generate_columns_to_use()
        df_ = make_complete_df(columns_to_use, y)
       
        if test:
            create_check_table_per_accotype(df_)
        else:
            df_prijzen_stacked = retrieve_prijzen()
            df = pd.merge(
                df_, df_prijzen_stacked, how="inner", on=["acco_type", "maand_str"]
            )
        
            df["omzet"] = df["in_house"] * df["price_per_night"]
         
            row = print_businessinfo(df, None, y)
            list.append(row)
            create_table_per_accotype_per_month(df, y)
            
        df_all_years = pd.concat([df_all_years, df_], axis=0).sort_values(by='datum')
    total_df = pd.DataFrame(list, columns=columns)
    st.write(total_df)
    
    st.subheader("Info over alle jaren")
    st.table(total_df)
    return df_all_years


def main():
    df_all_years = generate_and_show_info_all_years()
    

    df_all_years_non_grouped= make_date_columns(df_all_years)
    df_all_years_grouped = df_all_years_non_grouped.groupby(["datum"]).sum().reset_index()
    
    df_all_years_grouped = calculate_occupation(df_all_years_grouped)
    df_all_years_non_grouped = calculate_occupation(df_all_years_non_grouped)
    
   
    #df=df.groupby('Name').agg({'Missed':'sum', 'Credit':'sum','Grade':'mean'}).rename(columns=d)
    # df_all_years_grouped_maand_omzet = df_all_years_non_grouped.groupby(["jaar", "maand"]).sum().reset_index().pivot(index='jaar', columns='maand', values='omzet_eur')

    #print (df_all_years_grouped_maand_omzet)
    make_occopuation_graph(df_all_years_grouped)  
    make_occopuation_graph_per_acco(df_all_years_non_grouped)  
    most_checkins_out(df_all_years_grouped)
    make_booking_table(2022)

    
if __name__ == "__main__":
    main()
    #find_fill_color("B4")
    
