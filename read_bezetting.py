
import datetime as dt
from datetime import datetime, timedelta

import string
import pandas as pd

from openpyxl import load_workbook
import streamlit as st
from keys import * # secret file with the prices
import plotly.express as px
import plotly.graph_objects as go
import numpy as np

import re

test = True  # To test or not to test (to see if the fillcolors in the sheet are right.)
# https://github.com/onedrive/onedrive-sdk-python
# https://github.com/vgrem/Office365-REST-Python-Client#Working-with-OneDrive-API


excel_file = r"C:\Users\rcxsm\Downloads\bezetting2022.xlsm"
wb = load_workbook(excel_file, data_only=True)

excel_file_2023 =  r"C:\Users\rcxsm\Downloads\bezetting2023.xlsm"
wb_2023 = load_workbook(excel_file_2023, data_only=True)

what_to_do = st.sidebar.radio(
        "What to do",
        ('Checkin/out list', 'Analyse', "dekenanalyse","babypackanalyse","show info from bookingtable"))

        

if what_to_do == "Analyse":
    start_month, end_month = 3, 10
    
    (start_month,end_month) = st.sidebar.slider("Months (from/until (incl.))", 1, 12, (3,10))

    if start_month > end_month:
        st.warning("Make sure that the end month is not before the start month")
        st.stop()
   
    selection_list_accos_ = ['WAIKIKI' ,'BALI',  'KALAHARI1', 'KALAHARI2', 'SERENGETI XL', 'SERENGETI L', 'SAHARA']
    selection_list_accos = st.sidebar.multiselect("Welke acco's", selection_list_accos_,selection_list_accos_)
if what_to_do != 'Checkin/out list':
    year_ = [2019, 2021, 2022, 2023]
    year = st.sidebar.multiselect("Jaren", year_, [2021,2022, 2023])
sh_2022 = [wb["ALLES2022"]]
sh_2021 = [wb["ALLES2021"]]
sh_2019 = [wb["ALLES2019"]]
sh_2023 = [wb_2023["mrt april"], wb_2023["mei"], wb_2023["juni"], wb_2023["juli"], wb_2023["aug"], wb_2023["sept"], wb_2023["okt"]]
    # rij met de naam, startrij, eindrij
to_do_2023 = [
    [1, 4, 9],  # bal
    [11, 14, 23],  # wai
    [25, 28, 31],  # kal1
    [32,33, 36],  # kal2
    [37, 37, 45],  # kal1
    [46,46, 48],  # kal2
    [50, 53, 61],  # ser xl
    [62, 63, 70],  # ser L
    [71, 74, 87]] # navajo
to_do_2022 = [
    [1, 4, 9],  # bal
    [11, 14, 23],  # wai
    [25, 28, 40],  # kal1
    [41, 42, 48],  # kal2
    [50, 53, 61],  # ser xl
    [62, 63, 70],  # ser L
    [71, 75, 88],
]  # sahara
to_do_2021 = [
    [1, 4, 9],  # bal
    [11, 14, 23],  # wai
    [25, 28, 40],  # kal1
     [41, 42, 48],  # kal2
    [49, 52, 60],  # ser xl
    [62, 65, 86],
]  # sahara
to_do_2019 = [
    [3, 4, 9],  # bal 6
    [11, 12, 21],  # wai 10
    [23, 24, 36],  # kal1  13
    [38, 39, 47],  # sahara 9
    [49, 50, 58], #  ser xl 9
]  # 


def find_fill_color(cell):
    """Find fill color of a cell in 2022-tabblad
    Hulproutine, wordt niet aangeroepen in het script

    # dirty solution to find the fill color.
    # as solutions given here https://stackoverflow.com/questions/58429823/getting-excel-cell-background-themed-color-as-hex-with-openpyxl
    # dont work

    Args:
        cell (string): The cell you want to find the color from
    """

    val = sh_2022[cell].fill.start_color.rgb

    try:
        valx = val[0]
        valx = val
    except:
        valx = sh_2022[cell].fill.start_color.theme

    theme = sh_2022[cell].fill.start_color.theme
    tint = sh_2022[cell].fill.start_color.tint
    st.write(f"{valx = } {theme=} {tint=}")

    val = int(sh_2022[cell].fill.start_color.index, 16)
    st.write (val)
    #hex_color = "%06x" % (val && 0xFFFFFF)
    st.write(hex_color)
    

def retrieve_prijzen():
    """Retrieve the average price for an accomodation in a given month
    Returns:
        df: Table with the prices
    """    
    # sheet_id prijzen = in keys.py
    sheet_name_prijzen = "prijzen"
    #url_prijzen = f"https://docs.google.com/spreadsheets/d/{sheet_id_prijzen}/gviz/tq?tqx=out:csv&sheet={sheet_name_prijzen}"
    url_prijzen=r"C:\Users\rcxsm\Downloads\prijzen.csv"
    df_prijzen = pd.read_csv(url_prijzen, delimiter=",")
    # df_prijzen_stacked = df_prijzen.stack()
    df_prijzen_stacked = df_prijzen.melt(
        "acco_type", var_name="maand_int", value_name="price_per_night"
    )
    df_prijzen_stacked["maand_str"] = df_prijzen_stacked["maand_int"].astype(str)
    # .set_index('acco_type').stack().rename(columns={'price_per_night':'month'})
    return df_prijzen_stacked

def create_check_table_per_accotype(df,y):
    """Generate tables per accotype to see if the sheet is 100% correct (fill colors right*).
       The last column has to be zero

        * a booking starts with green
        * a booking end with cyaan (wissel) or red (checkout)
    Args:
        df (_type_): _description_
        y : year
    """

    list_of_accotypes_ = df.acco_type.unique()
   
    #list_of_accotypes = [list_of_accotypes_[4]]  # if you only one acco type
    year_ok=True
    for acco in list_of_accotypes_:
        df_acco = df[df["acco_type"] == acco].reset_index()
        df_acco = df_acco.groupby([df_acco["datum"]], sort=True).sum().reset_index()
        
        df_acco = df_acco.assign(bezet_saldo=None)
        df_acco.loc[0, "bezet_saldo"] = 0
        df_acco["bezet_theorie"] = (
            df_acco["geel"] + df_acco["wissel"] + df_acco["new_arrival"]
        )
        for i in range(1, len(df_acco)):
            df_acco.loc[i, "bezet_saldo"] = (
                df_acco.loc[i - 1, "bezet_saldo"]
                + df_acco.loc[i, "new_arrival"]
                - df_acco.loc[i, "vertrek_no_clean"]
                - df_acco.loc[i, "vertrek_clean"]
            )  
        df_acco["verschil_bezet"] = df_acco["bezet_theorie"] - df_acco["bezet_saldo"]
        df_acco_test = df_acco[df_acco["verschil_bezet"] != 0]
        if len(df_acco_test) ==0:
            print(f"{y} - {acco} OK")
        else:
            st.error(f"ERROR IN BOOKINGSSCHEMA {y} - {acco} ")
            st.write(df_acco_test)
            
            st.error("/ERROR ")
            year_ok=False
    if year_ok:
        print (f"{y} is okay")

            



def create_table_per_accotype_per_month(df, w, year):
    """ Maak een tabel voor een van de volgende zaken:
        omzet, verblijfsduur, bezetting, aantal overnachtingen, aantal boekingen
        per maand / per accotype in een bepaald jaar

    Args:
        df (_type_): _description_
        year (_type_): _description_
    """    
    # list_of_accotypes = df.acco_type.unique()
    # # list_of_accotypes = [list_of_accotypes_[8]] #if you only one acco type

    mm = ["month"]
    columns = [*mm, *selection_list_accos]
   
    months = [
        "januari",
        "februari",
        "maart",
        "april",
        "mei",
        "juni",
        "juli",
        "augustus",
        "september",
        "oktober",
        "november",
        "december",
    ]
    
    list_ = []
    for m in range(start_month, end_month + 1):
        rij = [months[m - 1]]
        for acco in selection_list_accos:

            df_acco = df[
                (df["acco_type"] == acco) & (df["maand_str"] == str(m))
            ].reset_index()

            in_tabel_te_plaatsen = round(generate_businessinfo(df_acco, w, year), 1)
            rij.append(in_tabel_te_plaatsen)

        list_.append(rij)
    
    table_per_accotype_per_month = pd.DataFrame(list_, columns=columns)
    
    if w == "omzet":   
        # calculate the totals for each column using a loop
        totals = {}
        for col in table_per_accotype_per_month.columns:
            if col != 'month':
                totals[col] = table_per_accotype_per_month[col].sum()
        totals['month'] = 'Total'
        
        # create a new DataFrame with the totals
        df_totals = pd.DataFrame(totals, index=[len(df)])

        # concatenate the original DataFrame and the totals DataFrame
        table_per_accotype_per_month = pd.concat([table_per_accotype_per_month, df_totals])
          
    return table_per_accotype_per_month
       
    


def generate_businessinfo(df_acco, what, year):
    """print and return the business intelligence voor een bepaald ajar
    Args:
        df_acco (_type_): _description_
        what (_type_): _description_

    Returns:
        _type_: _description_
    """
    aantal_boekingen =  df_acco["wissel"].sum() + df_acco["new_arrival"].sum()
   
    if (df_acco["aantal"].mean() * len(df_acco) - df_acco["out_of_order"].sum()) != 0:
        bezetting = round(
            (
                df_acco["geel"].sum()
                + df_acco["wissel"].sum()
                + df_acco["new_arrival"].sum()
            )
            / (
                (df_acco["aantal"].mean() * len(df_acco))
                - df_acco["out_of_order"].sum()
            )
            * 100,
            2,
        )
       
    else:
        bezetting = 0
    aantal_overnachtingen = df_acco["geel"].sum()  + df_acco["wissel"].sum()   + df_acco["new_arrival"].sum()
            
    if (df_acco["wissel"].sum() + df_acco["new_arrival"].sum()) != 0:
        # De verblijfsduur is vertekend als je het per maand, per acco bekijkt in rustige maanden, zie bijv. bali, september 2019 (maar 1 aankomst, maar mensen die nog vanuit augustus aanwezig zijn)
        verblijfsduur = round(
            (
                df_acco["geel"].sum()
                + df_acco["wissel"].sum()
                + df_acco["new_arrival"].sum()
            )
            / (df_acco["wissel"].sum() + df_acco["new_arrival"].sum()),
            2,
        )
        
    else:
        verblijfsduur = 0
    omzet = df_acco["omzet"].sum()

    # calculate number of accos for each type
    aantal_acco = 0
    if year == 2019:
        to_do = to_do_2019
        sh = sh_2019[0]
    if year == 2021:
        to_do = to_do_2021
        sh = sh_2021[0]
    if year == 2022:
        to_do = to_do_2022
        sh = sh_2022[0]
    if year == 2023:
        to_do = to_do_2023
        sh = sh_2023[0]
    for t in to_do:
        acco_type = str(sh["a" + str(t[0])].value)
        acco_type = replace_acco_types(acco_type)
        if acco_type in selection_list_accos:
            aantal_acco +=(t[2]-t[1]+1)


    
    if what == "omzet":
        to_return = omzet
    elif what == "verblijfsduur":
        to_return = verblijfsduur
    elif what == "bezetting":
    
        to_return = bezetting
    elif what == "aantal_boekingen":
        to_return  = aantal_boekingen
    elif what =='aantal_overnachtingen':
        to_return  = aantal_overnachtingen
    else:
        to_return = [year, omzet,aantal_acco, round((omzet/aantal_acco),2),round(bezetting,1),aantal_boekingen,round(verblijfsduur,1),aantal_overnachtingen,round(aantal_overnachtingen/aantal_acco,1)]
    return to_return

def generate_columns_to_use():
    """Generate a list with columns to use, eg. from A to ZZ

    Returns:
        _type_: _description_
    """
    alphabet = list(string.ascii_uppercase)
    alphabet_tot_h = [
        "",
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
    ]
    alphabet_to_use = []
    for alp in alphabet_tot_h:
        for alp2 in alphabet:
            alphabet_to_use.append(alp + alp2)
    alphabet_to_use = alphabet_to_use[1:]
    return alphabet_to_use

 
def add_row(list, acco_type,acco_number, guest_name, checkin_date, checkout_date, wissel_in, wissel_uit):
    """Add a row to the list.

    Args:
        list (list): the list where the new row has to be added
        acco_type (str): _description_
        acco_number (str): _description_
        guest_name (str): _description_
        checkin_date (str): _description_
        checkout_date (str): _description_

    Returns:
        list: _description_
    """    
    #number_of_days = datetime.strptime(checkout_date, "%Y-%m-%d") - datetime.strptime(checkin_date, "%Y-%m-%d")
    
    delta = datetime.strptime(checkout_date, "%Y-%m-%d").date() - datetime.strptime(checkin_date, "%Y-%m-%d").date()
    number_of_days = delta.days


    list.append([acco_type,acco_number, guest_name, checkin_date, checkout_date, wissel_in, wissel_uit, number_of_days])
    return list

def make_booking_table(year):
    """Generate a booking_tabel from the booking chart
    columns: [acco_type,acco_number, guest_name, checkin_date, checkout_date, number_of_days]

    Args:
        year (int): the year of which you want the booking table
    """

    columns_to_use = generate_columns_to_use()
    year__ = year
    to_do, sheets,sh_0 = select_to_do_and_sheet(year)
    list =[]
    
    for t in to_do:
        print (t)
        
        for r in range(t[1],t[2]+1):
            
            acco_number_cell = "a" + str(r)
            sh_titel = sheets[0]
            for sh in sheets:
                acco_type = str(sh_titel["a" + str(t[0])].value)
                acco_type = replace_acco_types(acco_type)
                acco_number = str(sh_titel[acco_number_cell].value)
            
                if acco_number == "kal  32m2  645":
                    acco_number = 645
                if acco_number ==    "kal 25m2 654":
                    acco_number = 654
                
                for c in columns_to_use[1:]:
                    cell_ = c +str(r)
                    datum_ = str(sh[c + "2"].value)
                    
                    try:
                        datum2 = datetime.strptime(datum_, "%Y-%m-%d %M:%H:%S")
                        datum = datetime.strftime(datum2, "%Y-%m-%d")
                        month = datum2.month
                        year = datum2.year
                    except:
                        datum = None
                        month = None
                        year = None

                    val = sh[cell_].fill.start_color.rgb
                    try:
                        valx = val[0]
                        valx = val
                    except:
                        valx = sh[cell_].fill.start_color.theme
                    if valx == 9:  # licht groen
                        checkin_date  = datum
                        guest_name = str(sh[c + str(r)].value)
                        wissel_in=False
                        
                    elif valx == "FFFF0000":  # rood
                        checkout_date  = datum
                        wissel_uit=False
                        list = add_row(list,acco_type,acco_number, guest_name, checkin_date, checkout_date, wissel_in, wissel_uit)
                    elif valx == "FF7030A0":  # paars
                        checkout_date  = datum
                        wissel_uit=False
                        list = add_row(list, acco_type,acco_number, guest_name, checkin_date, checkout_date, wissel_in, wissel_uit)
                    elif valx == 5:  # bruin
                        pass
                    elif valx == 0 or valx == 6:  # zwart of grijs
                        pass
                    elif valx == "FF00B0F0":  # lichtblauw / cyaan
                        checkout_date  = datum
                        wissel_uit = True
                        
                        
                        list = add_row(list,acco_type,acco_number, guest_name, checkin_date, checkout_date, wissel_in, wissel_uit)
                        checkin_date  = datum
                        guest_name = str(sh[c + str(r)].value)
                        wissel_in = True

                    elif valx == "FFFFFF00":  # geel / bezet
                        pass
                    #guest_name =""

                
    df = pd.DataFrame(
        list,
        columns=[
            "acco_type","acco_number", 
            "guest_name",
            "checkin_date",
            "checkout_date","wissel_in", "wissel_uit", "number_of_days"
        ],
    )


    return df 
def show_info_from_bookingtable(df, year):
    """Print the languages, linnen and babypacks

    Args:
        df (_type_): _description_
    """   
    st.subheader(f"Show info from bookingtable {year}")
    #df["number_of_days"] = df["number_of_days"].astype(string)
    # extract the language code from the text column using a regular expression
    df['language'] = df['guest_name'].str.extract(r'\b(du|en|fr|dk|gb|de|po|it|sp|ierl|be)\b', expand=False)
    # fill missing values with 'NL'
    df['language'] = df['language'].fillna('nl')
    df['guest_name'] = df['guest_name'].str.replace(r'\*1', 'x1')
    df['guest_name'] = df['guest_name'].str.replace(r'\*2', 'x2')
    df['x2'] = df['guest_name'].str.extract(r'\b(\d+)x2\b', expand=False)
    df['x1'] = df['guest_name'].str.extract(r'\b(\d+)x1\b', expand=False)
    
    df = df.fillna(0)
    df['babypack_old'] = df['guest_name'].str.contains('baby').astype(int)
    df['kst'] = df['guest_name'].str.contains('kst').astype(int)
    df['bb'] = df['guest_name'].str.contains(' bb').astype(int)
    df['babypack'] = np.where((df['babypack_old'] == 1) | (df['kst'] == 1) | (df['bb'] == 1), 1, 0)

    st.write(f"Total single {df['x1'].astype(int).sum()}")
    st.write(f"Total double {df['x2'].astype(int).sum()}")
    st.write(f"Total babypack {df['babypack'].sum()}")
    # get the frequency table for the 'values' column
    
    st.subheader(f"Talenverdeling in {year}")
    freq_table = df['language'].value_counts()
    st.write(freq_table)
    

def select_to_do_and_sheet(year):
    if year == 2019:
        to_do = to_do_2019
        sheets = sh_2019
    if year == 2021:
        to_do = to_do_2021
        sheets = sh_2021
    if year == 2022:
        to_do = to_do_2022
        sheets = sh_2022
    
    sh_0=None
    if year == 2023:
        to_do = to_do_2023
        sheets = sh_2023
        sh_0 = sh_2023[0] # de sheet waar de acconamen in 2023 uit worden gehaald
    return to_do,sheets, sh_0

def show_extra_info(year__, df):
    df['number_of_days'] = df['number_of_days'].dt.days.astype('int16')
    st.subheader(f"Distribution of length of stay in {year__}")
    
    st.write(f"Number of stays : {len(df)}")
    st.write (f"Number of days total : {df['number_of_days'].sum()}")
    st.write (f"Number of days min : {df['number_of_days'].min()}")
    st.write (f"Number of days max : {df['number_of_days'].max()}")
    st.write (f"Number of days average : {df['number_of_days'].mean()}")


    freq_tabel = df['number_of_days'].value_counts()
    st.write("Freq. number_of_days")
    fig = px.histogram(df, x="number_of_days")
    # plotly.offline.plot(fig)
    
    st.plotly_chart(fig, use_container_width=True)
   
  
def replace_acco_types(accotype_2023):
    # Define the string to be modified
   
    # Define a dictionary of the values to be replaced and their replacements
    replace_dict = {
        "Waikiki":"WAIKIKI",
        "BALI":"BALI",
        "Kalahari 32m2":"KALAHARI1",
        "Kalahari 25m2":"KALAHARI2",
        "Kalahari 32m2_":"KALAHARI1",
         
        "kal 25m2 654":"KALAHARI2",
        "Serengeti xl":"SERENGETI XL",
     
        "serengeti 5p":"SERENGETI L",
        "navajo":"SAHARA",
        "kal  32m2  645":"KALAHARI1",
    }
    # accotype_original = accotype_2023
    # for key, value in replace_dict.items():
    #     accotype_original = accotype_original.replace(key, value)

    # Create a regular expression pattern to match any of the keys in the dictionary
    pattern = re.compile("|".join(replace_dict.keys()))

    # Use the sub() method to replace all matches with their corresponding values
    accotype_original = pattern.sub(lambda m: replace_dict[m.group(0)], accotype_2023)

    return accotype_original

def make_mutation_df(year):
    """Generate the dataframe
        Columns: ['acco_type', 'aantal', 'datum',"month","year", "new_arrival","vertrek_no_clean", "vertrek_clean", "wissel", "geel"])

    Args:
        columns_to_use (list with strings): which columns to scrape, eg. from "A to ... ZZ "

    Returns:
        df: dataframe
    """
    columns_to_use = generate_columns_to_use()
    list_complete = []
    to_do, sheets,sh_0 = select_to_do_and_sheet(year)
   
    for ix,sh in enumerate(sheets):
        for a in columns_to_use:
            #       [rij met de naam, start, eind]'
            
            for t in to_do:
                
                if year == 2023:
                    acco_type = str(sh_0["a" + str(t[0])].value)
                else:
                    acco_type = str(sh["a" + str(t[0])].value)
                
                acco_type = replace_acco_types(acco_type)
               
                ii = []
                for x in range(t[1], t[2] + 1):
                    ii.append(a + str(x))
                bezet = 0
                aantal = t[2] - t[1] + 1
                vertrek_no_clean = 0
                vertrek_clean = 0
                vertrek_totaal = 0
                wissel = 0
                geel = 0
                out_of_order = 0
                new_arrival = 0
                try:
                    datum = str(sh[a + "2"].value)

                    datum2 = datetime.strptime(datum, "%Y-%m-%d %M:%H:%S")
                    datum3 = datetime.strftime(datum2, "%Y-%m-%d")
                    month = int(datum2.month)
                    year = datum2.year
                except:
                    datum3 = "STOP"
                if year == 2023:
                    # leave out the last day in every sheet (bv. 1 mei voor de april sheet)
                    stop_conditions = {
                        (5, 0): "STOP",
                        (6, 1): "STOP",
                        (7, 2): "STOP",
                        (8, 3): "STOP",
                        (9, 4): "STOP",
                        (10, 5): "STOP",
                    
                        
                    }
                    
                    datum3 = stop_conditions.get((month, ix), datum3)
                    
                if datum3 != "STOP":
                    for i in ii:
                        val = sh[i].fill.start_color.rgb
                        try:
                            valx = val[0]
                            valx = val
                        except:
                            valx = sh[i].fill.start_color.theme

                        if valx == "FFFF0000":  # rood
                            vertrek_clean += 1
                        elif valx == "FF7030A0":  # paars
                            vertrek_no_clean += 1
                        elif valx == 5:  # bruin
                            vertrek_totaal += 1
                        elif valx == 0 or valx == 6:  # zwart of grijs
                            out_of_order += 1
                        elif valx == "FF00B0F0":  # lichtblauw / cyaan
                            wissel += 1
                        elif valx == "FFFFFF00":  # geel / bezet
                            geel += 1
                        elif valx == 9:  # licht groen
                            new_arrival += 1

                    row = [
                        acco_type,
                        aantal,
                        datum3,
                        month,
                        year,
                        new_arrival,
                        vertrek_no_clean,
                        vertrek_clean,
                        wissel,
                        geel,
                        out_of_order,
                    ]
                    list_complete.append(row)
    df_mutation = pd.DataFrame(
        list_complete,
        columns=[
            "acco_type",
            "aantal",
            "datum",
            "month",
            "year",
            "new_arrival",
            "vertrek_no_clean",
            "vertrek_clean",
            "wissel",
            "geel",
            "out_of_order",
        ],
    )
    df_mutation["in_house"] = df_mutation["geel"] + df_mutation["new_arrival"] + df_mutation["wissel"]
    df_mutation["maand_str"] = df_mutation["month"].astype(str)
    df_mutation = df_mutation[(df_mutation["month"] >= start_month) & (df_mutation["month"] <= end_month)]
    
    
    df_mutation = df_mutation[df_mutation["acco_type"].isin(selection_list_accos)]
   
    return df_mutation

def make_date_columns(df):
   

    df['datum'] = pd.to_datetime(df.datum, format='%Y-%m-%d')
    df["jaar"] = df["datum"].dt.strftime("%Y")
    df["maand"] = df["datum"].dt.strftime("%m").astype(str).str.zfill(2)
    df["dag"] = df["datum"].dt.strftime("%d").astype(str).str.zfill(2)
    df["maand_dag"] = df["maand"] + "-" + df["dag"]
    df["dag_maand"] = df["dag"] + "-" + df["maand"]
    df["datum_str"] = df["datum"].astype(str)
    df["datum_"] = pd.to_datetime(df["maand_dag"], format="%m-%d")
    # convert the date column to a datetime data type
    return df


def  make_occopuation_graph_per_acco(df_):
    """_summary_

    GIVES AN ERROR ValueError: Index contains duplicate entries, cannot reshape


    Args:
        df_ (_type_): _description_

    """ 
      
    for y in year:
        df_all_years_grouped = df_[df_["jaar"] == str(y)]
        print("718")
        print(df_all_years_grouped)
        df_all_years_grouped = df_all_years_grouped.groupby(df_all_years_grouped["acco_type","datum"]).sum().reset_index()

        data = []
        fig = go.Figure()
        print("722")
        print(df_all_years_grouped)
        df_all_years_grouped = df_all_years_grouped.sort_values(by='datum') 
        
        
        df_all_years_grouped = df_all_years_grouped.drop("datum_", axis=1)
  
        df_all_years_pivot_a = df_all_years_grouped.reset_index().pivot(index=['datum'], columns='acco_type', values='occupation')
        df_all_years_pivot_a = df_all_years_pivot_a.fillna(0).reset_index()
       
        width, opacity = 1,1
        column_names = df_all_years_pivot_a.columns[1:].tolist()
     
        #for a in selection_list_accos:
        for a in column_names:
            try:
                # niet alle accomodaties zijn gebruikt in alle jaren
                points = go.Scatter(
                                name = a,
                                x=df_all_years_pivot_a["datum"],
                                y=df_all_years_pivot_a[a],
                                line=dict(width=width),
                                opacity=opacity,
                                mode="lines",
                            )
                    
                data.append(points)
            except:
                pass
        layout = go.Layout(
                    yaxis=dict(title=f"Bezetting (%)"),
                    title=f"Bezetting per acco type in {y}",
                )
        fig = go.Figure(data=data, layout=layout)
        fig.update_layout(xaxis=dict(tickformat="%d-%m"))
        #fig.show()
        #plotly.offline.plot(fig)
        st.plotly_chart(fig, use_container_width=True)
        

def  make_occopuation_graph(df_all_years_grouped):
    data = []
    fig = go.Figure()
    df_all_years_grouped = df_all_years_grouped.sort_values(by='datum_') 

   
    df_all_years_grouped["jaar"] = df_all_years_grouped["jaar"].astype(str)
    
    
    df_all_years_pivot = df_all_years_grouped.pivot(index=['maand_dag','datum_'], columns='jaar', values='occupation').reset_index()

    for y in year:
        y_=str(y)
        if y_ == "2023":
            width = 2
            opacity = 1
        else:
            width = 0.7
            opacity = 0.8

        points = go.Scatter(
                        name = y_,
                        x=df_all_years_pivot["datum_"],
                        y=df_all_years_pivot[y_],
                        line=dict(width=width),
                        opacity=opacity,
                        mode="lines",
                    )
            
        data.append(points)
    layout = go.Layout(
                yaxis=dict(title=f"Bezetting (%)"),
                title="Bezetting",
            )
    fig = go.Figure(data=data, layout=layout)
    fig.update_layout(xaxis=dict(tickformat="%d-%m"))
    #fig.show()
    #plotly.offline.plot(fig)
    st.plotly_chart(fig, use_container_width=True)
    # st.write(df_all_years_pivot)
    # save_df(df_all_years_pivot,"bezetting_per_jaar.csv")
    
def save_df(df, name):
    """  _ _ _ """
    name_ =  name + ".csv"
    compression_opts = dict(method=None, archive_name=name_)
    df.to_csv(name_, index=False, compression=compression_opts)

    print("--- Saving " + name_ + " ---")


def most_checkins_out(df_all_years_grouped):

    for y in year:
        df_all_years_grouped_one_year = df_all_years_grouped[df_all_years_grouped["jaar"]==str(y)]
       
        df_checkouts = df_all_years_grouped_one_year.sort_values(by='checkouts',ascending=False)
        checkouts = df_checkouts[["datum", "checkouts"]]
        df_checkins = df_all_years_grouped_one_year.sort_values(by='checkins', ascending=False) 
        checkins = df_checkins[["datum", "checkins"]]
        st.subheader (y)
        
        col1,col2=st.columns(2)
        with col1:
            st.subheader("Checkouts")
            fig2a = px.bar(df_all_years_grouped_one_year, x='datum', y='checkouts', title=f'Number of checkouts over Time in {y}')
            st.plotly_chart(fig2a, use_container_width=True)
        
            st.write (checkouts.head(10))
            fig = px.histogram(df_checkouts, x="checkouts", title=y)
            st.plotly_chart(fig, use_container_width=True)
        with col2:
            st.subheader("Checkins")
            fig2b = px.bar(df_all_years_grouped_one_year, x='datum', y='checkins', title=f'Number of checkins over Time in {y}')
            st.plotly_chart(fig2b, use_container_width=True)
    
            st.write (checkins.head(10))
            fig = px.histogram(df_checkins, x="checkins", title=y)
            st.plotly_chart(fig, use_container_width=True)

def calculate_occupation(df):
    
    df["checkins"] = df["wissel"] + df["new_arrival"]
    df["checkouts"] = df["wissel"] + df["vertrek_no_clean"]+ df["vertrek_clean"]
    df["occupation"] = round(100*(df["new_arrival"]+df["wissel"]+df["geel"]) /
                                                (df["aantal"]-df["out_of_order"]),2)
   
    df= make_date_columns(df)

    
    return df


def generate_info_all_years():
    df_mutations_all_years =  pd.DataFrame()
    info,columns = [], ["year","omzet_eur","aantal_acco","omzet_per_acco","bezetting_%","aantal_boekingen","gem_verblijfsduur","aantal_overnachtingen","aantal_overnachtingen_per_acco"]
    for y in year:
        
        df_mutation_year = make_mutation_df(y)
        
        if test:
            print (f"------------{y}-----------")
            create_check_table_per_accotype(df_mutation_year,y)
        
        df_prijzen_stacked = retrieve_prijzen()
        
        df_mutations_met_omzet = pd.merge(
            df_mutation_year, df_prijzen_stacked, how="inner", on=["acco_type", "maand_str"]
        )
        df_mutations_met_omzet["omzet"] = df_mutations_met_omzet["in_house"] * df_mutations_met_omzet["price_per_night"]
      
        row = generate_businessinfo(df_mutations_met_omzet, None, y)
        info.append(row)   
        df_mutations_all_years = pd.concat([df_mutations_all_years, df_mutations_met_omzet], axis=0).sort_values(by='datum')
    df_info_per_jaar = pd.DataFrame(info, columns=columns)
    

    return df_mutations_all_years, df_info_per_jaar

def show_info_all_years(df_mutations_met_omzet, info_per_jaar_df):
    df_mutations_all_years =  pd.DataFrame()
    for y in year:
        

        row = generate_businessinfo(df_mutations_met_omzet, None, y)

        # to_return = [year, omzet,aantal_acco, round((omzet/aantal_acco),2),round(bezetting,1),aantal_boekingen,round(verblijfsduur,1),aantal_overnachtingen,round(aantal_overnachtingen/aantal_acco,1)]
        st.subheader(f"Info over {y}")
        st.write(f"Aantal boekingen {row[5]}")
        st.write(f"Bezetting {row[4]} %")
        st.write(f"Gemiddeld verblijfsduur {row[6]}")
        st.write(f"Totale omzet (indicatie): {row[1]}")
        show_detailed_info_per_year = True
        if show_detailed_info_per_year:
            what = ["omzet", "verblijfsduur", "bezetting","aantal_overnachtingen","aantal_boekingen"]
            for w in what:
                st.write()
                st.subheader(f"{w} -- {y} ---")
                table_per_accotype_per_month = create_table_per_accotype_per_month(df_mutations_met_omzet,w, y)
                st.write(table_per_accotype_per_month.transpose().astype(str))
      
   
    st.subheader("Info over alle jaren - total_df")
    st.table(info_per_jaar_df)

    return df_mutations_all_years

def babypackanalyse(df, y):
    st.header(f"Babypack analyse - {y}")
     
    df['babypack_old'] = df['guest_name'].str.contains('baby').astype(int)
    df['kst'] = df['guest_name'].str.contains('kst').astype(int)
    df['bb'] = df['guest_name'].str.contains(' bb').astype(int)
    df['babypack'] = np.where((df['babypack_old'] == 1) | (df['kst'] == 1) | (df['bb'] == 1), 1, 0)

    #NOG AANPASSEN
    df_ = df[df['babypack'] ==1]
    st.write("Bookings with babypack")
    st.write(df_)
    # convert date columns to datetime format
    df['checkin_date'] = pd.to_datetime(df['checkin_date'])
    df['checkout_date'] = pd.to_datetime(df['checkout_date'])

    # create a date range to cover all possible dates
    date_range = pd.date_range(start=f'{y}-04-8', end=f'{y}-09-30', freq='D')

    # initialize a dictionary to store the totals for each date
    totals = {}

    # iterate over the date range and calculate the total guests with babypacks for each date
    for date in date_range:
        mask = ((df['checkin_date'] <= date) & (df['checkout_date'] > date) & df['babypack'] ==1 )
        total = df.loc[mask, 'guest_name'].count()
        totals[date] = total

    # create a dataframe from the dictionary and convert the index to a date column
    df_babypacks = pd.DataFrame.from_dict(totals, orient='index', columns=['total_babypacks'])
    df_babypacks.reset_index(inplace=True)
    df_babypacks.rename(columns={'index': 'date'}, inplace=True)
    df_babypacks['date'] = pd.to_datetime(df_babypacks['date'])

    fig = px.line(df_babypacks, x='date', y='total_babypacks', title=f'Number of Babypacks over Time in {y}')
    st.plotly_chart(fig, use_container_width=True)
    freq_tabel = df_babypacks['total_babypacks'].value_counts()
    st.write("Aantal dagen dat x babypacks in gebruik zijn")
    st.write(freq_tabel)
    st.write(f"Maximum aantal totaal aantal babypacks {df_babypacks['total_babypacks'].max()}")
   


def deken_analyse(df_bookingtable, year):
    """ Calculate how many blankets we need on stock, based on the maximum occupation of the sahara's in 2022
        With the help of ChatGPT
    Args:
        df_bookingtable (_type_): _description_
    """    
    
    df_bookingtable = df_bookingtable[df_bookingtable["acco_type"] == "SAHARA"] 
    
    # filter rows where col2 ends with 'xp' and the character before 'p' is a number
    df_bookingtable_filtered = df_bookingtable[df_bookingtable['guest_name'].str.match('.*[0-9]p$')]
    # extract the number before 'xp' in each row and store it in a new column
    df_bookingtable_filtered['number_of_guests'] = df_bookingtable_filtered['guest_name'].str.extract('(\d+)p', expand=False)

    df_bookingtable_filtered['number_of_guests'] = df_bookingtable_filtered['number_of_guests'].astype(int)
    
    freq_tabel = df_bookingtable_filtered['number_of_guests'].value_counts()
    fig = px.histogram(df_bookingtable_filtered, x="number_of_guests")
    # plotly.offline.plot(fig)
    st.write ("Freq table number_of_guests (per boeking)")
    st.plotly_chart(fig, use_container_width=True)
    st.write ("Freqeuncy table number of guests")
    st.write (freq_tabel)
    st.write(f"Gemiddelde gezinsgrootte  {round(df_bookingtable_filtered['number_of_guests'].mean(),2)}")

        
    # convert date columns to datetime format
    df_bookingtable_filtered['checkin_date'] = pd.to_datetime(df_bookingtable_filtered['checkin_date'])
    df_bookingtable_filtered['checkout_date'] = pd.to_datetime(df_bookingtable_filtered['checkout_date'])

    # create a date range to cover all possible dates
    date_range = pd.date_range(start='2022-06-15', end='2022-09-13', freq='D')

    # initialize a dictionary to store the totals for each date
    totals = {}

    # iterate over the date range and calculate the total guests for each date
    for date in date_range:
        mask = (df_bookingtable_filtered['checkin_date'] <= date) & (df_bookingtable_filtered['checkout_date'] > date)
        total = df_bookingtable_filtered.loc[mask, 'number_of_guests'].sum()
        totals[date] = total

    # convert the dictionary to a dataframe and print it
    totals_df_bookingtable_filtered = pd.DataFrame.from_dict(totals, orient='index', columns=['total_guests']) #.to_string()
   
    freq_tabel = totals_df_bookingtable_filtered['total_guests'].value_counts()
    fig = px.histogram(totals_df_bookingtable_filtered, x="total_guests", nbins=int((totals_df_bookingtable_filtered['total_guests'].max() - totals_df_bookingtable_filtered['total_guests'].min()) / 1))
    # plotly.offline.plot(fig)
    st.write ("Freq table total  number_of_guests in Sahara start='2022-06-15', end='2022-09-13' ")
    st.plotly_chart(fig, use_container_width=True)
    st.write(f"Maximum aantal totaal aantal gasten {totals_df_bookingtable_filtered['total_guests'].max()}")
    st.write(f"Maximum aantal dekens per tent {round( totals_df_bookingtable_filtered['total_guests'].max()/14,1)}")


    # get list of unique dates
    #dates = df_bookingtable_filtered['checkin_date'].unique()
    dates = pd.date_range(start='2022-06-15', end='2022-09-13', freq='D')
    # create empty dataframe to store results
    result_df = pd.DataFrame()

    # loop over dates and compute total number of guests in each room
    for date in dates:
        df_filtered = df_bookingtable_filtered[(df_bookingtable_filtered['checkin_date'] <= date) & (df_bookingtable_filtered['checkout_date'] > date)]
       
        
        #result_df[date] = df_grouped
        
        # df_grouped.columns = ['acco_number', 'number_of_guests', "date"]
        df_filtered['date'] = date
        result_df = pd.concat([result_df, df_filtered])
    # transpose the result dataframe
    
    #df_grouped = result_df.groupby(['acco_number'])['number_of_guests'].sum().reset_index()
    
    result_df["acco_number"] = result_df["acco_number"].astype(str).str.zfill(2)

    df_pivot = result_df.pivot_table(index='date', columns='acco_number', values='number_of_guests',  aggfunc='sum')
   
    st.write("Bezetting per tent per dag")
    # display result
    st.write(df_pivot)
def occupation_per_accomodation(year):

    """Which accomodation is occupied the most?
    """    
    df_bookingtable = make_booking_table(year)

    df_grouped = df_bookingtable.groupby(["acco_type","acco_number"], sort=True).sum().reset_index()
    df_grouped["occupation"] = round((df_grouped["number_of_days"] / (30+31+30+31+31+30) *100),1)  #183 nachten, maart laten we buiten beschouwing
    st.subheader(f"Occupation per accomodation {year}")
    st.write (df_grouped)




def checkin_outlist(df_bookingtable, date_to_check_):
   
    # CSS to inject contained in a string
    hide_table_row_index = """
            <style>
            thead tr th:first-child {display:none}
            tbody th {display:none}
            </style>
            """

    # Inject CSS with Markdown
    st.markdown(hide_table_row_index, unsafe_allow_html=True)
   
    st.subheader(f"Checkouts {date_to_check_}")
    df_bookingtable_out = df_bookingtable[df_bookingtable["checkout_date"] == date_to_check_]
    df_bookingtable_out = df_bookingtable_out[["acco_number","guest_name", "wissel_uit"]]

    st.table(df_bookingtable_out)

    st.subheader(f"Checkins {date_to_check_}")
    df_bookingtable_in = df_bookingtable[df_bookingtable["checkin_date"] == date_to_check_]
    df_bookingtable_in = df_bookingtable_in[["acco_number","guest_name", "checkout_date"]]
    
    st.table(df_bookingtable_in)
    return 

def main():
    if what_to_do == 'Analyse':
        df_mutations_all_years,df_info_per_jaar = generate_info_all_years()
        analyse(df_mutations_all_years,df_info_per_jaar)
    elif what_to_do == 'Checkin/out list':
        date_to_check_ = st.sidebar.date_input("Date to check").strftime('%Y-%m-%d')# .strptime("%Y-%m-%d")
        year_  = dt.datetime.strptime(date_to_check_, "%Y-%m-%d").date().year
        df_bookingtable = make_booking_table(year_)
        
        checkin_outlist(df_bookingtable,date_to_check_ ) 
    else:
        for y in year:
            df_bookingtable = make_booking_table(y)
        
           
            if what_to_do == "dekenanalyse":
                deken_analyse(df_bookingtable,y)
            elif what_to_do == "babypackanalyse" and y==2022:
                babypackanalyse(df_bookingtable, y)
            elif what_to_do == "show info from bookingtable":
                show_info_from_bookingtable(df_bookingtable, y)
    
  
def analyse(df_mutations_all_years,df_info_per_jaar):
    
    show_info_all_years(df_mutations_all_years,df_info_per_jaar)
    
    df_all_years_non_grouped= make_date_columns(df_mutations_all_years)
  
    df_all_years_grouped = df_all_years_non_grouped.groupby(["datum"]).sum().reset_index()
    
    df_all_years_grouped = calculate_occupation(df_all_years_grouped)

    st.subheader("Omzet per maand")
    df_all_years_non_grouped = calculate_occupation(df_all_years_non_grouped)
    df_all_years_grouped_maand_omzet = df_all_years_non_grouped.groupby(["jaar", "maand"]).sum().reset_index().pivot(index='jaar', columns='maand', values='omzet')
    
    st.write(df_all_years_grouped_maand_omzet)
 
    make_occopuation_graph(df_all_years_grouped)  

    # gives error
    #make_occopuation_graph_per_acco(df_all_years_non_grouped)  
    most_checkins_out(df_all_years_grouped)


    # find_fill_color("H80")
    
    for y in year:
        st.header(" ----------------- Extra info {y} -------------")
        occupation_per_accomodation(y)
       
          
        

if __name__ == "__main__":

  
    main()
   

