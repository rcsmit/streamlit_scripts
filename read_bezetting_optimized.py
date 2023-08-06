import re
import datetime as dt
from datetime import datetime #, timedelta
# import datetime

import string
import pandas as pd
import time
from openpyxl import load_workbook
import streamlit as st
# from keys import *  # secret file with the prices
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
import sys
import platform
import calendar

def clear_cache():
    """Clears the cache
    """
    if st.button("Clear All"):
        # Clear values from *all* all in-memory and on-disk data caches:
        # i.e. clear values from both square and cube
        st.cache_data.clear()
        st.write("Cache is cleared")
    st.stop()

def find_fill_color():

    """Find fill color of a cell in 2023-tabblad

    # dirty solution to find the fill color.
    # as solutions given here https://stackoverflow.com/questions/58429823/getting-excel-cell-background-themed-color-as-hex-with-openpyxl
    # dont work

    Args:
        cell (string): The cell you want to find the color from
    """
    if platform.processor() != "":
        sheet_name = st.sidebar.text_input("Sheetname", "KLADBLOK")
        cell = st.sidebar.text_input("Cell", "A1")
        excel_file_2023 = r"C:\Users\rcxsm\Downloads\bezetting2023a.xlsm"
        wb_2023 = load_workbook(excel_file_2023, data_only=True)

        sh_2023 = wb_2023[sheet_name]

        val = sh_2023[cell].fill.start_color.rgb

        try:
            valx = val[0]
            valx = val
        except Exception:
            valx = sh_2023[cell].fill.start_color.theme

        theme = sh_2023[cell].fill.start_color.theme
        tint = sh_2023[cell].fill.start_color.tint
        val = int(sh_2023[cell].fill.start_color.index, 16)

        st.write(f"File = {excel_file_2023}")
        st.write(f"Cel = {cell}")
        st.write(f"{valx = } {theme=} {tint=}")

        st.write(f"{val = }")
        # hex_color = "%06x" % (val && 0xFFFFFF)
        # st.write(hex_color)
    else:
        st.error("Available only on local systems")
        st.stop()

@st.cache_data
def retrieve_prijzen():
    """Retrieve the average price for an accomodation in a given month
    Returns:
        df: Table with the prices
    """
    # sheet_id prijzen = in keys.py
    # sheet_name_prijzen = "prijzen"
    # url_prijzen = f"https://docs.google.com/spreadsheets/d/{sheet_id_prijzen}/gviz/tq?tqx=out:csv&sheet={sheet_name_prijzen}"
    if platform.processor() != "":
    
        url_prijzen = r"C:\Users\rcxsm\Downloads\prijzen.csv"
    else:
        url_prijzen = "https://raw.githubusercontent.com/rcsmit/streamlit_scripts/main/input/prijzen_dummy.csv"
    df_prijzen = pd.read_csv(url_prijzen, delimiter=",")
    # df_prijzen_stacked = df_prijzen.stack()
    df_prijzen_stacked = df_prijzen.melt(
        "acco_type", var_name="month_int", value_name="price_per_night"
    )
    df_prijzen_stacked["month_str"] = df_prijzen_stacked["month_int"].astype(str)
    # .set_index('acco_type').stack().rename(columns={'price_per_night':'month'})
    return df_prijzen_stacked

def create_check_table_per_accotype(df, y):
    """Generate tables per accotype to see if the sheet is 100% correct (fill colors right*).
       The last column has to be zero

        * a booking starts with green
        * a booking end with cyaan (back_to_back) or red (checkout)
    Args:
        df (_type_): _description_
        y : year
    """
    
    list_of_accotypes_ = df.acco_type.unique()
  
    st.write(f"Checking {y}")
    # list_of_accotypes = [list_of_accotypes_[4]]  # if you only one acco type
    year_ok = True
    for acco in list_of_accotypes_:
        
        df_acco = df[df["acco_type"] == acco].reset_index()
        df_acco = df_acco.groupby([df_acco["date"]], sort=True)[["geel","back_to_back","new_arrival", "vertrek_no_clean", "vertrek_clean"]].sum().reset_index()
        #df_acco = df_acco.groupby(["date"]["geel","back_to_back","new_arrival", "vertrek_no_clean", "vertrek_clean"], sort=True).sum().reset_index()
        df_acco = df_acco.assign(bezet_saldo=None)
        df_acco.loc[0, "bezet_saldo"] = 0
        df_acco["bezet_theorie"] = (
            df_acco["geel"] + df_acco["back_to_back"] + df_acco["new_arrival"]
        )
        for i in range(1, len(df_acco)):
            if y == "2023":
                # paars is ook bezet ipv vertrek clean
                df_acco.loc[i, "bezet_saldo"] = (
                    df_acco.loc[i - 1, "bezet_saldo"]
                    + df_acco.loc[i, "new_arrival"]
                    - df_acco.loc[i, "vertrek_no_clean"]
                    - df_acco.loc[i, "vertrek_clean"]
                )
            else:
                df_acco.loc[i, "bezet_saldo"] = (
                    df_acco.loc[i - 1, "bezet_saldo"]
                    + df_acco.loc[i, "new_arrival"]
                    - df_acco.loc[i, "vertrek_no_clean"]
                    - df_acco.loc[i, "vertrek_clean"]
                )
        df_acco["verschil_bezet"] = df_acco["bezet_theorie"] - df_acco["bezet_saldo"]
        
        
        df_acco_test = df_acco[df_acco["verschil_bezet"] != 0]
        if len(df_acco_test) == 0:
            st.write(f":white_check_mark: {y} - {acco} OK")
            with st.expander("DF"):
                st.write (df_acco)
           
        else:
            st.error(f":heavy_exclamation_mark: ERROR IN BOOKINGSSCHEMA {y} - {acco} ")
            with st.expander("Complete DF"):
                st.write(df_acco)
            st.write(df_acco_test)

            st.error("/ERROR ")
            year_ok = False
    if year_ok:
        st.write(f":white_check_mark: **{y} is okay**")

def generate_businessinfo(df_mutations):
    """print and return the business intelligence
    Args:
        df_mutations (_type_): df met info over accos
        df_mutation : df met de mutaties
        y : year
        
    Returns:
        _type_: _description_
    """
    st.write(df_mutations)
    aantal_boekingen = df_mutations["back_to_back"].sum() + df_mutations["new_arrival"].sum()

    if (df_mutations["aantal"].mean() * len(df_mutations) - df_mutations["out_of_order"].sum()) != 0:
        occupation = round(
            (
                df_mutations["geel"].sum()
                + df_mutations["back_to_back"].sum()
                + df_mutations["new_arrival"].sum()
            )
            / (
                (df_mutations["aantal"].mean() * len(df_mutations))
                - df_mutations["out_of_order"].sum()
            )
            * 100,
            2,
        )

    else:
        occupation = 0
    aantal_overnachtingen = (
        df_mutations["geel"].sum() + df_mutations["back_to_back"].sum() + df_mutations["new_arrival"].sum()
    )

    if (df_mutations["back_to_back"].sum() + df_mutations["new_arrival"].sum()) != 0:
        # De verblijfsduur is vertekend als je het per month, per acco bekijkt in rustige maanden, zie bijv. bali, september 2019 (maar 1 aankomst, maar mensen die nog vanuit augustus aanwezig zijn)
        verblijfsduur = round(
            (
                df_mutations["geel"].sum()
                + df_mutations["back_to_back"].sum()
                + df_mutations["new_arrival"].sum()
            )
            / (df_mutations["back_to_back"].sum() + df_mutations["new_arrival"].sum()),
            2,
        )

    else:
        verblijfsduur = 0
    omzet = df_mutations["omzet"].sum()

    st.write(f"{aantal_boekingen=}")
    st.write(f"{occupation=}")
    st.write(f"{aantal_overnachtingen=}")
    st.write(f"{verblijfsduur=}")
    st.write(f"{omzet=}")
    # calculate number of accos for each type
   

def generate_columns_to_use():
    """Generate a list with columns to use, eg. from A to ZZ

    Returns:
        _type_: _description_
    """
    alphabet = list(string.ascii_uppercase)
    alphabet_tot_h = [
        "",
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
    ]
    alphabet_to_use = []
    for alp in alphabet_tot_h:
        for alp2 in alphabet:
            alphabet_to_use.append(alp + alp2)
    alphabet_to_use = alphabet_to_use[1:]
    return alphabet_to_use

def add_row(
    list,
    acco_type,
    acco_number,
    guest_name,
    checkin_date,
    checkout_date,
    back_to_back_in,
    back_to_back_out,
):
    """Add a row to the bookingslist. Called from make_booking_table()

    Args:
        list (list): the list where the new row has to be added
        acco_type (str): _description_
        acco_number (str): _description_
        guest_name (str): _description_
        checkin_date (str): _description_
        checkout_date (str): _description_

    Returns:
        list: _description_
    """
    number_of_days = datetime.strptime(checkout_date, "%Y-%m-%d") - datetime.strptime(checkin_date, "%Y-%m-%d")
    try:
        delta = (
            datetime.strptime(checkout_date, "%Y-%m-%d").date()
            - datetime.strptime(checkin_date, "%Y-%m-%d").date()
        )
        number_of_days = delta.days
    except:
        st.write(f" {acco_number= }, {guest_name=}, {checkin_date=}, {checkout_date=},")
        number_of_days = 0

    list.append(
        [
            acco_type,
            acco_number,
            guest_name,
            checkin_date,
            checkout_date,
            back_to_back_in,
            back_to_back_out,
            number_of_days,
        ]
    )
    return list

# @st.cache_data()
def make_booking_table(wb_2023):
    """Generate a booking_tabel from the booking chart
    columns: [acco_type,acco_number, guest_name, checkin_date, checkout_date, number_of_days]

    Args:
        year (int): the year of which you want the booking table
    """

    columns_to_use = generate_columns_to_use()
    list = []
    # placeholder_what = st.empty()
    acco_type, acco_temp = None, None
    # placeholder_progress = st.empty()
    for year in [2019,2021,2022,2023]:
        year_temp = year
        to_do, sheets= select_to_do_and_sheet(wb_2023, year)
        
        for i,t in enumerate(to_do):
            progress = int(i/len(to_do)*100)
            for r in range(t[1], t[2] + 1):

                first_guest = False
                guest_name = None
                acco_number_cell = "a" + str(r)
                sh_titel = sheets[0]
                for sh in sheets:
                    if acco_type != acco_temp:
                        print(f"Making booking table   {year_temp} | {acco_type}" )
                        acco_temp = acco_type
                    # placeholder_progress.progress(progress)
           
                    acco_type = str(sh_titel["a" + str(t[0])].value)
                    acco_type = replace_acco_types(acco_type)
                    acco_number = str(sh_titel[acco_number_cell].value)

                    if acco_number == "kal  32m2  645":
                        acco_number = 645
                    if acco_number == "kal 25m2 654":
                        acco_number = 654
                    if acco_number == "25m2 641":
                        acco_number = 641
                    if acco_number == "25m2 655":
                        acco_number = 655

                    for c in columns_to_use[1:]:
                        cell_ = c + str(r)
                        date_ = str(sh[c + "2"].value)

                        try:
                            date2 = datetime.strptime(date_, "%Y-%m-%d %M:%H:%S")
                            date = datetime.strftime(date2, "%Y-%m-%d")
                            month = date2.month
                            year = date2.year
                        except Exception:
                            date = None
                            month = None
                            year = None

                        val = sh[cell_].fill.start_color.rgb
                        try:
                            valx = val[0]
                            valx = val
                        except Exception:
                            valx = sh[cell_].fill.start_color.theme
                        # if valx != "00000000":
                        #     print (f"{sh} {r}  {c} {valx} {sh[c + str(r)].value}")
                        if valx == 9 or valx == "FF70AD47":  # licht groen
                            first_guest = True
                            checkin_date = date
                            # if (guest_name == "" or guest_name == None) and first_guest == True:
                            #     guest_name = "ERROR"
                            # else:
                            guest_name = str(sh[c + str(r)].value)
                            #st.write(f"--{r}--{sh}---{val} --{cell_}- {guest_name}")
                            back_to_back_in = False

                        elif valx == "FFFF0000":  # rood
                            checkout_date = date
                            back_to_back_out = False
                            list = add_row(
                                list,
                                acco_type,
                                acco_number,
                                guest_name,
                                checkin_date,
                                checkout_date,
                                back_to_back_in,
                                back_to_back_out,
                            )
                            guest_name = "_"
                        elif valx == "FF7030A0":  # paars
                            checkout_date = date
                            back_to_back_out = False
                            list = add_row(
                                list,
                                acco_type,
                                acco_number,
                                guest_name,
                                checkin_date,
                                checkout_date,
                                back_to_back_in,
                                back_to_back_out,
                            )
                        elif valx == 5:  # bruin
                            pass
                        elif valx == 0 or valx == 6:  # zwart of grijs
                            pass
                        elif valx == "FF00B0F0" or valx=="FFFFC000":  # lichtblauw / cyaan - dark yellow for non paid resa
                            checkout_date = date
                            back_to_back_out = True

                            list = add_row(
                                list,
                                acco_type,
                                acco_number,
                                guest_name,
                                checkin_date,
                                checkout_date,
                                back_to_back_in,
                                back_to_back_out,
                            )
                            checkin_date = date
                            guest_name = str(sh[c + str(r)].value)
                            back_to_back_in = True

                        elif valx == "FFFFFF00":  # geel / bezet
                            pass
                        # guest_name =""

    df = pd.DataFrame(
        list,
        columns=[
            "acco_type",
            "acco_number",
            "guest_name",
            "checkin_date",
            "checkout_date",
            "back_to_back_in",
            "back_to_back_out",
            "number_of_days",
        ],
    )
    df = extract_info(df)
    df = add_extra_linnen(df)
    df = make_date_columns(df)
    # placeholder_what.empty()
    # placeholder_progress.empty()
    return df

def extract_info(df):
    """Extracts nationality, guest name, linnen, babypack and bookingnumber and puts in a new column

    Args:
        df (_type_): _description_

    Returns:
        _type_: _description_
    """
    # df["number_of_days"] = df["number_of_days"].astype(string)
    # extract the language code from the text column using a regular expression
    df["guest_name"] = df["guest_name"].str.replace(r'^de\s', 'de_', regex=True)
    df["guest_name"] = df["guest_name"].str.replace(r'^van de\s', 'van de_', regex=True)
    df["guest_name"] = df["guest_name"].str.replace("sp+", "s_peeltuin+", regex=True) #geeft een fout bij "nispen sp+"
    df["guest_name"] = df["guest_name"].str.replace(r"\bsp\+\b", "s_peeltuin+", regex=True) #geeft een fout bij "nispen sp+"
    

    df["language"] = df["guest_name"].str.extract(
        r"\b(du|de|en|fr|dm|dk|gb|de|uk|po|it|ph|lx|lux|ch|sp|ierl|be)\b", expand=False
    )
    df['language'] = df['language'].replace({'gb':'en','du': 'de',  'dm': 'dk'})

    # fill missing values with 'NL'
    df["language"] = df["language"].fillna("nl")

    df["guest_name_booking"] = df["guest_name"].str.split().str[0]
    # Extract the text and store it in a new column
    # df['guest_name_booking'] = df['guest_name'].apply(lambda x: re.match(r'^(.*?)(?=\b(?:\d+|du|en|fr|dk|gb|de|po|it|sp|ierl|be|)\b)', x).group(1))
    df["guest_name_booking"] = df["guest_name"].str.extract(r"^([^0-9]+)")
    # Remove specific endings from the extracted text
    endings = ["du", "en", "fr", "dk", "dm","uk", "gb", "de", "po", "it", "sp", "ierl", "be"]
    # Replace values in the column

    df["guest_name_booking"] = df["guest_name_booking"].str.replace(
        rf'({"|".join(endings)})$', "", regex=True
    )
    df["guest_name_booking"] = df["guest_name_booking"].str.replace("s_peeltuin+", "sp", regex=True) #geeft een fout bij "nispen sp+"

    # LINNEN
    df["guest_name"] = df["guest_name"].str.replace(r"\*1", "x1", regex=True)
    df["guest_name"] = df["guest_name"].str.replace(r"\*2", "x2", regex=True)
    df["dbl_linnen"] = df["guest_name"].str.extract(r"\b(\d+)x2\b", expand=False).fillna(0).astype(int)
    df["sng_linnen"] = df["guest_name"].str.extract(r"\b(\d+)x1\b", expand=False).fillna(0).astype(int)

    df = df.fillna(0)
    df["babypack_old"] = df["guest_name"].str.contains("baby").astype(int)
    df["kst"] = df["guest_name"].str.contains("kst").astype(int)
    df["bb"] = df["guest_name"].str.contains(" bb").astype(int)
    df["babypack"] = np.where(
        (df["babypack_old"] == 1) | (df["kst"] == 1) | (df["bb"] == 1), 1, 0
    )

    df["booking_number"] = df["guest_name"].str.extract(r"(\d+)").fillna(0).astype(int)
    # Function to format the numbers
    def format_number(number):
        formatted_number = format(number, ",")
        return formatted_number.replace(",", "")

    # Apply the formatting function to the 'Number' column
    df["booking_number"] = df["booking_number"].apply(format_number)
    df["guest_name"] = df["guest_name"].str.replace(r'^de_\s', 'de', regex=True)
    df["guest_name"] = df["guest_name"].str.replace(r'^van de_\s', 'van de', regex=True)
    return df

def add_extra_linnen(booking_table):

    import pandas as pd

    # Assuming you have two dataframes: bookingtable and extra_linnen
    # with a common column 'reservation nr' and a column 'single linnen'

    # Example dataframes
  
    data_extra_linnen = {'reservation nr': [2, 3, 4],
                        'single linnen': [2, 3, 1]}

    bookingtable = pd.DataFrame(data_bookingtable)
    extra_linnen = pd.DataFrame(data_extra_linnen)

    # Merge dataframes on 'reservation nr' and update 'single linnen' column
   # Merge dataframes on 'reservation nr' and update the columns
    merged_df = bookingtable.merge(extra_linnen, on='booking_number', how='left')
    columns_to_update = ['dbl_linnen', 'sng_linnen', 'kst', 'bb']

    for col in columns_to_update:
        merged_df[col + '_x'] += merged_df[col + '_y']
        merged_df.drop(columns=[col + '_y'], inplace=True)
        merged_df.rename(columns={col + '_x': col}, inplace=True)

    # Print the updated dataframe
    print(merged_df)






    # Print the updated dataframe
    print(merged_df)
    return booking_table


def graph_distribution_nationalities(df, y):
    """Show a graph with the distribution of nationalities in time

    Args:
        df (df): booking table
        y : year, used in the graph titles
    """    
    df =df.copy()
    # Convert the 'checkin_date' and 'checkout_date' columns to datetime format
    df['checkin_date'] = pd.to_datetime(df['checkin_date'])
    df['checkout_date'] = pd.to_datetime(df['checkout_date'])

    # Generate a date range from the minimum checkin_date to the maximum checkout_date
    date_range = pd.date_range(df['checkin_date'].min(), df['checkout_date'].max())

    # Calculate the proportion of each language on each day
    language_proportions = pd.DataFrame()

    for date in date_range:
        guests_on_date = df[(df['checkin_date'] <= date) & (df['checkout_date'] > date)]
        guests_on_date = guests_on_date.copy()  
        guests_on_date["date_language"] = date

        # language_counts = guests_on_date['language'].value_counts(normalize=True)
        language_proportions = pd.concat([language_proportions,guests_on_date], ignore_index=True)

    
    language_proportions["count"] =1
    # Create the pivot table
    df_pivot = pd.pivot_table(language_proportions, index='date_language', columns="language", values="count", aggfunc='sum')
    
    # normalize the pivot table to ensure each row sums to 100%
    pivot_table_normalized = df_pivot.div(df_pivot.sum(axis=1), axis=0) * 100
    # Round the values to one decimal place
    pivot_table_normalized = pivot_table_normalized.round(1).fillna(0)
    # Reorder the columns to have NL at the bottom, followed by DE, and then the rest of the languages
    pivot_table_normalized = pivot_table_normalized[['nl', 'de', "be"] + [col for col in pivot_table_normalized.columns if col not in ['nl', 'de', "be"]]]
   
    # Create a line graph using Plotly
    fig = go.Figure()

    for column in pivot_table_normalized.columns:
        fig.add_trace(go.Scatter(
            x=pivot_table_normalized.index,
            y=pivot_table_normalized[column],
            name=column
        ))

    fig.update_layout(
        title=f'Proportions of Languages over Time  in {y}',
        xaxis_title='Date',
        yaxis_title='Proportion (%)',
    )
    st.plotly_chart(fig)

        
    
    # Create a stacked bar graph using Plotly
    fig = go.Figure()

    for column in pivot_table_normalized.columns:
        fig.add_trace(go.Bar(
            x=pivot_table_normalized.index,
            y=pivot_table_normalized[column],
            name=column,
            offsetgroup=column,
        ))

    fig.update_layout(
        title=f'Stacked Proportions of Languages over Time in {y}',
        xaxis_title='Date',
        yaxis_title='Proportion (%)',
        barmode='stack',
    )
    st.plotly_chart(fig)

    df_pivot_acco = pd.pivot_table(language_proportions, index='acco_type', columns="language", values="count", aggfunc='sum').fillna(0)
    st.write("df_pivot_acco")
    st.write(df_pivot_acco)


    pivot_table_acco_normalized_row = df_pivot_acco.div(df_pivot_acco.sum(axis=1), axis=0) * 100
 
    pivot_table_acco_normalized_row = pivot_table_acco_normalized_row.round(1)
    st.write("pivot_table_acco_normalized_row")
    st.write(pivot_table_acco_normalized_row)


    pivot_table_acco_normalized_col = df_pivot_acco.div(df_pivot_acco.sum(axis=0), axis=1) * 100
    pivot_table_acco_normalized_col = pivot_table_acco_normalized_col.round(1)
    st.write("pivot_table_acco_normalized_col")
    st.write(pivot_table_acco_normalized_col)

def show_info_from_bookingtable(df, year):
    """Print the languages, linnen and babypacks

    Args:
        df (_type_): _description_
    """
    st.subheader(f"Show info from bookingtable {year}")
    

    st.write(f"Total single {df['sng_linnen'].astype(int).sum()}")
    st.write(f"Total double {df['dbl_linnen'].astype(int).sum()}")
    st.write(f"Total babypack {df['babypack'].sum()}")
    # get the frequency table for the 'values' column

    st.subheader(f"Distribution of languages in {year}")
    freq_table = df["language"].value_counts()
    st.write(freq_table)
    st.write(f"Total number of bookings :{len(df)}")
    
    graph_distribution_nationalities(df, year)

def select_to_do_and_sheet(wb_2023, year):

    """Select which rows to do for a given year
        Called from make_booking_table and make_mutation_df

    Returns:
        _type_: _description_
    """    

    sh_2022 = [wb_2023["ALLES2022"]]
    sh_2021 = [wb_2023["ALLES2021"]]
    sh_2019 = [wb_2023["ALLES2019"]]
    sh_2023 = [
        wb_2023["mrt april"],
        wb_2023["mei"],
        wb_2023["juni"],
        wb_2023["juli"],
        wb_2023["aug"],
        wb_2023["sept"],
        wb_2023["okt"],
        wb_2023["nov"],
    ]
    # rij met de naam, startrij, eindrij
    to_do_2023 = [
        [1, 4, 9],  # bal
        [11, 14, 23],  # wai
        [25, 28, 31],  # kal1
        [32, 33, 36],  # kal2
        [37, 37, 45],  # kal1
        [46, 46, 48],  # kal2
        [50, 53, 61],  # ser xl
        [62, 63, 70],  # ser L
        [71, 74, 87],
    ]  # navajo
    to_do_2022 = [
        [1, 4, 9],  # bal
        [11, 14, 23],  # wai
        [25, 28, 40],  # kal1
        [41, 42, 48],  # kal2
        [50, 53, 61],  # ser xl
        [62, 63, 70],  # ser L
        [71, 75, 88],
    ]  # sahara
    to_do_2021 = [
        [1, 4, 9],  # bal
        [11, 14, 23],  # wai
        [25, 28, 40],  # kal1
        [41, 42, 48],  # kal2
        [49, 52, 60],  # ser xl
        [62, 65, 86],
    ]  # sahara
    to_do_2019 = [
        [3, 4, 9],  # bal 6
        [11, 12, 21],  # wai 10
        [23, 24, 36],  # kal1  13
        [38, 39, 47],  # sahara 9
        [49, 50, 58],  #  ser xl 9
    ]  #
    
    if year == 2019:
        to_do = to_do_2019
        sheets = sh_2019
        #sh_0 = sh_2019[0]
    elif year == 2021:
        to_do = to_do_2021
        sheets = sh_2021
        #sh_0 = sh_2019[0]
    elif year == 2022:
        to_do = to_do_2022
        sheets = sh_2022
        #sh_0 = sh_2022[0]
    
    elif year == 2023:
        to_do = to_do_2023
        sheets = sh_2023
        #sh_0 = sh_2023[0]  # de sheet waar de acconamen in 2023 uit worden gehaald
    else:
        st.error(f"ERROR IN YEAR year = {year}")

    return to_do, sheets #, sh_0

def show_info_number_of_days(df,y):
    #df["number_of_days"] = df["number_of_days"].dt.days.astype("int16")
    st.subheader(f"Distribution of length of stay in {y}")

    st.write(f"Number of stays : {len(df)}")
    st.write(f"Number of days total : {df['number_of_days'].sum()}")
    st.write(f"Number of days min : {df['number_of_days'].min()}")
    st.write(f"Number of days max : {df['number_of_days'].max()}")
    st.write(f"Number of days average : {df['number_of_days'].mean()}")

    freq_tabel = df["number_of_days"].value_counts()
    st.write("Freq. number_of_days")
    fig = px.histogram(df, x="number_of_days")
    # plotly.offline.plot(fig)

    st.plotly_chart(fig, use_container_width=True)

def replace_acco_types(accotype_2023):
    # Define the string to be modified

    # Define a dictionary of the values to be replaced and their replacements
    replace_dict = {
        "Waikiki": "WAIKIKI",
        "BALI": "BALI",
        "Kalahari 32m2": "KALAHARI1",
        "Kalahari 25m2": "KALAHARI2",

        "kalahari 25m2": "KALAHARI2",
        "kal 2            654": "KALAHARI2",
        
        "kal  1           645": "KALAHARI1",
        "kalahari 2 ": "KALAHARI2",
        "Kalahari 1": "KALAHARI1",
        "kal 32m2 645": "KALAHARI1",
       
        "kalahari 25m2": "KALAHARI2",
        "Kalahari 32m2_": "KALAHARI1",
        "kal 25m2 654": "KALAHARI2",
        "Serengeti xl": "SERENGETI XL",
        "serengeti 5p": "SERENGETI L",
        "navajo": "SAHARA",
        "kal  32m2  645": "KALAHARI1",
        "Kalahari 32m2": "KALAHARI1",
        "Kalahari 25m2": "KALAHARI2",
        "kal  32m2  645": "KALAHARI1",
        "kal 25m2 654": "KALAHARI2",
        "25m2          641": "KALAHARI2",
    }
    # accotype_original = accotype_2023
    # for key, value in replace_dict.items():
    #     accotype_original = accotype_original.replace(key, value)

    # Create a regular expression pattern to match any of the keys in the dictionary
    pattern = re.compile("|".join(replace_dict.keys()))

    # Use the sub() method to replace all matches with their corresponding values
    accotype_original = pattern.sub(lambda m: replace_dict[m.group(0)], accotype_2023)

    return accotype_original

# @st.cache_data
def make_mutation_df(wb_2023):
    """Generate the dataframe
        Columns: ['acco_type', 'aantal', 'date',"month","year", "new_arrival","vertrek_no_clean", "vertrek_clean", "back_to_back", "geel"])

    Args:
        columns_to_use (list with strings): which columns to scrape, eg. from "A to ... ZZ "

    Returns:
        df: dataframe
    """
    columns_to_use = generate_columns_to_use()
    list_complete = []
    # placeholder_what = st.empty()
    # placeholder_progress = st.empty()
    for year in [2019,2021,2022,2023]:

        to_do, sheets = select_to_do_and_sheet(wb_2023, year)
        
        for ix, sh in enumerate(sheets):
            print(f"Reading {year} - {sh}  [{ix+1}/{len(sheets)}]")
            for i,a in enumerate(columns_to_use):

                # progress = int(i/len(columns_to_use)*100)
                # placeholder_progress.progress(progress)
                #       [rij met de naam, start, eind]'
                
                for t in to_do:

                    #if year == 2023:
                    acco_type = str(sh["a" + str(t[0])].value)
                   
                    if acco_type == "kal  32m2  645" or acco_type =="637":
                        acco_type = "KALAHARI1"
                    if acco_type == "25m2          641" or  acco_type == "kal 25m2 654" or acco_type =="641":
                        acco_type = "KALAHARI2"
                    #else:
                    #    acco_type = str(sh["a" + str(t[0])].value)
                    # if t[0] ==32 :
                    #     print (f".{acco_type}.")
                    acco_type = replace_acco_types(acco_type)

                    ii = []
                    for x in range(t[1], t[2] + 1):
                        ii.append(a + str(x))
                    bezet = 0
                    aantal = t[2] - t[1] + 1
                    
                    vertrek_no_clean = 0
                    vertrek_clean = 0
                    vertrek_totaal = 0
                    back_to_back = 0
                    geel = 0
                    out_of_order = 0
                    new_arrival = 0
                    try:
                        date = str(sh[a + "2"].value)
                        date2 = datetime.strptime(date, "%Y-%m-%d %M:%H:%S")
                        date3 = datetime.strftime(date2, "%Y-%m-%d")
                        month = int(date2.month)
                        year = date2.year
                    except Exception:
                        date3 = "STOP"
                        month = 0

                    if year == 2023:
                        # leave out the last day in every sheet (bv. 1 mei voor de april sheet)
                        stop_conditions = {
                            (5, 0): "STOP",
                            (6, 1): "STOP",
                            (7, 2): "STOP",
                            (8, 3): "STOP",
                            (9, 4): "STOP",
                            (10, 5): "STOP",
                        }

                        date3 = stop_conditions.get((month, ix), date3)

                    if date3 != "STOP":
                        for i in ii:
                            val = sh[i].fill.start_color.rgb
                            try:
                                valx = val[0]
                                valx = val
                            except Exception:
                                valx = sh[i].fill.start_color.theme

                            if valx == "FFFF0000":  # rood
                                vertrek_clean += 1
                            elif valx == "FF7030A0":  # paars
                                vertrek_no_clean += 1
                            elif valx == 5:  # bruin
                                vertrek_totaal += 1
                            elif valx == 0 or valx == 6:  # zwart of grijs
                                out_of_order += 1
                            elif valx == "FF00B0F0"  or valx=="FFFFC000":  # lichtblauw / cyaan - dark yellow for non paid resa
                                back_to_back += 1
                            elif valx == "FFFFFF00":  # geel / bezet
                                geel += 1
                            elif valx == 9 or valx == "FF70AD47":  # licht groen
                                new_arrival += 1

                        row = [
                            acco_type,
                            aantal,
                            date3,
                            month,
                            year,
                            new_arrival,
                            vertrek_no_clean,
                            vertrek_clean,
                            back_to_back,
                            geel,
                            out_of_order,
                        ]
                        list_complete.append(row)
    df_mutation = pd.DataFrame(
        list_complete,
        columns=[
            "acco_type",
            "aantal",
            "date",
            "month",
            "year",
            "new_arrival",
            "vertrek_no_clean",
            "vertrek_clean",
            "back_to_back",
            "geel",
            "out_of_order",
        ],
    )
    df_mutation["in_house"] = (
        df_mutation["geel"] + df_mutation["new_arrival"] + df_mutation["back_to_back"]
    )
    df_mutation["month_str"] = df_mutation["month"].astype(str)
    # df_mutation = df_mutation[
    #     (df_mutation["month"] >= start_month) & (df_mutation["month"] <= end_month)
    # ]

    #df_mutation = df_mutation[df_mutation["acco_type"].isin(selection_list_accos)]
    #st.write(f"{year=}, {start_month=}, {end_month=} {selection_list_accos=}")
    df_mutation = make_date_columns(df_mutation)

    # add turnover info
    df_prijzen_stacked = retrieve_prijzen()

    df_mutations_met_omzet = pd.merge(
            df_mutation,
            df_prijzen_stacked,
            how="inner",
            on=["acco_type", "month_str"],
        )
    df_mutations_met_omzet["omzet"] = (
            df_mutations_met_omzet["in_house"]
            * df_mutations_met_omzet["price_per_night"]
        )
    # placeholder_what.empty()
    # placeholder_progress.empty()
    return df_mutations_met_omzet

def make_date_columns(df):
    datefields = ["date", "checkin_date", "Arrival Date"]
    for d in datefields:
        try:
            df["date"] = pd.to_datetime(df[d], format="%Y-%m-%d")
        except Exception:
            pass
    df["year"] = df["date"].dt.strftime("%Y")
    df["year_int"] = df["date"].dt.strftime("%Y").astype(int)
    df["month"] = df["date"].dt.strftime("%m").astype(str).str.zfill(2)
    df["day"] = df["date"].dt.strftime("%d").astype(str).str.zfill(2)
    df["month_day"] = df["month"] + "-" + df["day"]
    df["day_month"] = df["day"] + "-" + df["month"]
    df["date_str"] = df["date"].astype(str)
    df["date_"] = pd.to_datetime(df["month_day"], format="%m-%d")
    # convert the date column to a datetime data type
    return df

def make_occupuation_graph(df___):
    """_summary_

    GIVES AN ERROR ValueError: Index contains duplicate entries, cannot reshape

    Args:
        df_ (_type_): _description_

    """
    st.subheader("Occupation all accomodationtypes, all years")
    df_grouped_by_date_year = df___.groupby(["year","date", "day_month"])[["aantal", "in_house"]].sum().sort_values(by="date").reset_index()
    df_grouped_by_date_year["occupation"] = round(df_grouped_by_date_year["in_house"]/ df_grouped_by_date_year["aantal"]*100,1)
       
  
    df_grouped_by_date_year["day_month_dt"] = pd.to_datetime(df_grouped_by_date_year["day_month"], format="%d-%m")
    pivot_table = df_grouped_by_date_year.pivot_table(index="day_month_dt", columns="year", values="occupation")

    figz = go.Figure()
    for column in pivot_table.columns:
        figz.add_trace(go.Scatter(x=pivot_table.index, y=pivot_table[column], name=str(column)))
    
    figz.update_layout(xaxis_tickformat="%d-%b")  # Display only the day and month on x-axis

    st.plotly_chart(figz, use_container_width=True)

    years = [2019,2021,2022,2023]
    for y in years:
        st.subheader(y)
        df_mutations_year = df___[df___["year"] == str(y)]
        df_grouped_by_date = df_mutations_year.groupby("date")[["aantal", "in_house"]].sum().sort_values(by="date").reset_index()
        df_grouped_by_date["occupation"] = round(df_grouped_by_date["in_house"]/ df_grouped_by_date["aantal"]*100,1)
        
        
        df__ = df_mutations_year.groupby(["date", "acco_type"])[["aantal", "in_house"]].sum().sort_values(by="date").reset_index()
        df__["occupation"] = round(df__["in_house"]/ df__["aantal"]*100,1)
        
        
        width, opacity = 1, 1

        fig1 = go.Figure(data=go.Scatter(x=df_grouped_by_date["date"], y=df_grouped_by_date["occupation"], mode="lines"))
        st.plotly_chart(fig1, use_container_width=True)
        
    
        fig = go.Figure()
    
        df_all_years_pivot_a = df__.pivot(
            index=["date"], columns="acco_type", values="occupation"
        ).fillna(0).reset_index()
    
        
        column_names = df_all_years_pivot_a.columns[1:].tolist()
        data =[]
        # for a in selection_list_accos:
        for a in column_names:
            try:
                # niet alle accomodaties zijn gebruikt in alle jaren
                points = go.Scatter(
                    name=a,
                    x=df_all_years_pivot_a["date"],
                    y=df_all_years_pivot_a[a],
                    line=dict(width=width),
                    opacity=opacity,
                    mode="lines",
                )

                data.append(points)
            except Exception:
                pass
        layout = go.Layout(
            yaxis=dict(title=f"Occupation (%)"),
            title=f"Occupation per acco type ",
        )
        fig = go.Figure(data=data, layout=layout)
        fig.update_layout(xaxis=dict(tickformat="%d-%m"))
        # fig.show()
        # plotly.offline.plot(fig)
        st.plotly_chart(fig, use_container_width=True)

def save_df(df, name):
    """_ _ _"""
    name_ = name + ".csv"
    compression_opts = dict(method=None, archive_name=name_)
    df.to_csv(name_, index=False, compression=compression_opts)

    print("--- Saving " + name_ + " ---")

def check_mutation_table(df_mutation):
   
    for y in ["2023"]:
        df_mutation_year = df_mutation[df_mutation["year"] == y]
        
        print(f"------------{y}-----------")
        create_check_table_per_accotype(df_mutation_year, y)

def generate_info_all_years(df_mutation, years, selection_list_accos):
   
    info, columns = [], [
        "year",
        "omzet_eur",
        "aantal_acco",
        "omzet_per_acco",
        "occupation_%",
        "aantal_boekingen",
        "gem_verblijfsduur",
        "aantal_overnachtingen",
        "aantal_overnachtingen_per_acco",
    ]
    for y in years:

        row = generate_businessinfo(df_mutation, y, selection_list_accos)
        info.append(row)
       
    df_info_per_year = pd.DataFrame(info, columns=columns)

    return  df_info_per_year

def babypackanalyse(df, y):
    st.subheader(f"Babypack analyse - {y}")
    df =df.copy()
    df["babypack_old"] = df["guest_name"].str.contains("baby").astype(int)
    df["kst"] = df["guest_name"].str.contains("kst").astype(int)
    df["bb"] = df["guest_name"].str.contains(" bb").astype(int)
    df["babypack"] = np.where(
        (df["babypack_old"] >= 1) | (df["kst"] >= 1) | (df["bb"] >= 1), 1, 0
    )

    # NOG AANPASSEN
    df_ = df[df["babypack"] == 1]
    st.write("Bookings with babypack")
    st.write(df_)
    # convert date columns to datetime format
    df["checkin_date"] = pd.to_datetime(df["checkin_date"])
    df["checkout_date"] = pd.to_datetime(df["checkout_date"])

    # create a date range to cover all possible dates
    date_range = pd.date_range(start=f"{y}-04-8", end=f"{y}-09-30", freq="D")

    # initialize a dictionary to store the totals for each date
    totals = {}

    # iterate over the date range and calculate the total guests with babypacks for each date
    for date in date_range:
        mask = (df["checkin_date"] <= date) & (df["checkout_date"] > date) & df[
            "babypack"
        ] == 1
        total = df.loc[mask, "guest_name"].count()
        totals[date] = total

    # create a dataframe from the dictionary and convert the index to a date column
    df_babypacks = pd.DataFrame.from_dict(
        totals, orient="index", columns=["total_babypacks"]
    )
    df_babypacks.reset_index(inplace=True)
    df_babypacks.rename(columns={"index": "date"}, inplace=True)
    df_babypacks["date"] = pd.to_datetime(df_babypacks["date"])

    fig = px.line(
        df_babypacks,
        x="date",
        y="total_babypacks",
        title=f"Number of Babypacks over Time in {y}",
    )
    st.plotly_chart(fig, use_container_width=True)
    freq_tabel = df_babypacks["total_babypacks"].value_counts()
    st.write("Aantal days dat x babypacks in gebruik zijn")
    st.write(freq_tabel)
    st.write(
        f"Maximum aantal totaal aantal babypacks {df_babypacks['total_babypacks'].max()}"
    )

    if y == 2023:
        date_to_check_ = st.sidebar.date_input("Date to check")
    
        desired_date = dt.datetime.combine(date_to_check_, dt.datetime.min.time())
        guests_with_babybed = df[(df['checkin_date'] <= desired_date) & (df['checkout_date'] > desired_date) & (df['bb'] >= 1)]
        guests_with_highchair = df[(df['checkin_date'] <= desired_date) & (df['checkout_date'] > desired_date) & (df['kst'] >= 1)]
        st.subheader("Guests with babybed at {date_to_check} ({len(guests_with_babybed)})")
        st.write(guests_with_babybed)
        st.subheader("Guests with highchair at {date_to_check} ({len(guests_with_highchair)})")
        st.write(guests_with_highchair)

def deken_analyse(df_bookingtable, year):
    """Calculate how many blankets we need on stock, based on the maximum occupation of the sahara's in 2022
        With the help of ChatGPT

        Gives strange results for 2019 and 2021

    Args:
        df_bookingtable (_type_): _description_
    """
    st.subheader(f"Dekenanalyse {year}")
    df_bookingtable = df_bookingtable[(df_bookingtable["acco_type"] == "SAHARA")]
    if len (df_bookingtable) > 0:
        # filter rows where col2 ends with 'xp' and the character before 'p' is a number
        # df_bookingtable_filtered = df_bookingtable[
        #     df_bookingtable["guest_name"].str.match(".*[0-9]p$")
        # ]

        df_bookingtable_filtered = df_bookingtable[df_bookingtable["guest_name"].str.contains(r"\d+p")]

        df_bookingtable_filtered = df_bookingtable_filtered.copy()
        st.write (df_bookingtable_filtered)
        # extract the number before 'xp' in each row and store it in a new column
        # df_bookingtable_filtered["number_of_guests"] = df_bookingtable_filtered[
        #     "guest_name"
        # ].str.extract("(\d+)p", expand=False)
        df_bookingtable_filtered["number_of_guests"] = df_bookingtable_filtered["guest_name"].str.extract("(\d+)(?=p)", expand=False).astype(int)
        #df_bookingtable_filtered["number_of_guests"] = df_bookingtable_filtered["guest_name"].str.extract("(\d+)p", expand=False).astype(int)
      
        with st.expander("df_bookingtable_filtered"):
            st.write(df_bookingtable_filtered)
        freq_tabel = df_bookingtable_filtered["number_of_guests"].value_counts()
        fig = px.histogram(df_bookingtable_filtered, x="number_of_guests")
        # plotly.offline.plot(fig)
        st.write("Freq table number_of_guests (per boeking)")
        st.plotly_chart(fig, use_container_width=True)
        st.write("Freqeuncy table number of guests")
        st.write(freq_tabel)
        st.write(
            f"Gemiddelde gezinsgrootte  {round(df_bookingtable_filtered['number_of_guests'].mean(),2)}"
        )

        # convert date columns to datetime format
        df_bookingtable_filtered["checkin_date"] = pd.to_datetime(
            df_bookingtable_filtered["checkin_date"]
        )
        df_bookingtable_filtered["checkout_date"] = pd.to_datetime(
            df_bookingtable_filtered["checkout_date"]
        )

        # create a date range to cover all possible dates
        date_range = pd.date_range(start=f"{year}-04-15", end=f"{year}-09-30", freq="D")

        # initialize a dictionary to store the totals for each date
        totals = {}

        # iterate over the date range and calculate the total guests for each date
        for date in date_range:
            mask = (df_bookingtable_filtered["checkin_date"] <= date) & (
                df_bookingtable_filtered["checkout_date"] > date
            )
            total = df_bookingtable_filtered.loc[mask, "number_of_guests"].sum()
            totals[date] = total

        # convert the dictionary to a dataframe and print it
        totals_df_bookingtable_filtered = pd.DataFrame.from_dict(
            totals, orient="index", columns=["total_guests"]
        )  # .to_string()

        freq_tabel = totals_df_bookingtable_filtered["total_guests"].value_counts()
        fig = px.histogram(
            totals_df_bookingtable_filtered,
            x="total_guests",
            nbins=int(
                (
                    totals_df_bookingtable_filtered["total_guests"].max()
                    - totals_df_bookingtable_filtered["total_guests"].min()
                )
                / 1
            ),
        )
        # plotly.offline.plot(fig)
        st.write(
            f"Freq table total  number_of_guests in Sahara start='{year}-04-15', end='{year}-09-30' "
        )
        st.plotly_chart(fig, use_container_width=True)
        st.write(
            f"Maximum aantal totaal aantal gasten {totals_df_bookingtable_filtered['total_guests'].max()}"
        )
        st.write(
            f"Maximum aantal dekens per tent {round( totals_df_bookingtable_filtered['total_guests'].max()/14,1)}"
        )

        # get list of unique dates
        # dates = df_bookingtable_filtered['checkin_date'].unique()
        #dates = pd.date_range(start=f"{year}-06-15", end=f"{year}-09-13", freq="D")
        # create empty dataframe to store results
        result_df = pd.DataFrame()

        # loop over dates and compute total number of guests in each room

        for date in date_range:
            df_filtered = df_bookingtable_filtered[
                (df_bookingtable_filtered["checkin_date"] <= date)
                & (df_bookingtable_filtered["checkout_date"] > date)
            ].copy()

            # Set the value using .loc
            df_filtered["date"] = date

            # Concatenate the filtered DataFrame to the result_df
            result_df = pd.concat([result_df, df_filtered], ignore_index=True)

        result_df["acco_number"] = result_df["acco_number"].astype(str).str.zfill(2)
        #result_df["acco_number"] = result_df["acco_number"].astype(int)
        df_pivot = result_df.pivot_table(
            index="date", columns="acco_number", values="number_of_guests", aggfunc="sum"
        ).fillna(0)

        st.write("occupation per tent per day")
        # display result
        st.write(df_pivot)

        # Calculate the total number of guests per day
        df_pivot['Total'] = df_pivot.sum(axis=1)

        # Create the graph using Plotly Express
        fig = px.line(df_pivot, x=df_pivot.index, y='Total', title='Total Number of Guests per Day')

        # Show the graph
        st.plotly_chart(fig)

        # Create a line graph using Plotly
        fig = go.Figure()

        for column in df_pivot.columns[:-1]:
            fig.add_trace(go.Bar(
                x=df_pivot.index,
                y=df_pivot[column],
                name=column,
                offsetgroup = column
            ))

        fig.update_layout(
            title=f'Blankets over Time  in {year}',
            xaxis_title='Date',
            yaxis_title='Blankets',
            barmode='stack',
        )
        st.plotly_chart(fig)
    else:
        st.error(f"Geen boekingen in Sahara in year {year}")

def checkin_outlist(df_bookingtable, date_to_check_, current_date_day_of_week):

    # CSS to inject contained in a string
    hide_table_row_index = """
            <style>
            thead tr th:first-child {display:none}
            tbody th {display:none}
            </style>
            """

    # Inject CSS with Markdown
    st.markdown(hide_table_row_index, unsafe_allow_html=True)

    df_bookingtable_out = df_bookingtable[
        df_bookingtable["checkout_date"] == date_to_check_
    ]
    df_bookingtable_out = df_bookingtable_out[
        ["acco_number", "guest_name", "back_to_back_out"]
    ]
    st.subheader(
        f"Checkouts {current_date_day_of_week} {date_to_check_} | aantal = {len(df_bookingtable_out)} | b2b = {len(df_bookingtable_out[df_bookingtable_out['back_to_back_out'] == True])} "
    )
    if len(df_bookingtable_out) > 0:
        st.table(df_bookingtable_out)

    df_bookingtable_in = df_bookingtable[
        df_bookingtable["checkin_date"] == date_to_check_
    ]
    df_bookingtable_in = df_bookingtable_in[
        ["acco_number", "guest_name", "checkout_date",  "back_to_back_in"]
    ]
    st.subheader(
        f"Checkins {current_date_day_of_week} {date_to_check_} | aantal = {len(df_bookingtable_in)} | b2b = {len(df_bookingtable_in[df_bookingtable_in['back_to_back_in'] == True])}"
    )
    if len(df_bookingtable_in) > 0:
        st.table(df_bookingtable_in)
    st.markdown("<hr>", unsafe_allow_html=True)
    nr_in = len(df_bookingtable_in)
    nr_out = len(df_bookingtable_out)
    back_to_back =len(df_bookingtable_out[df_bookingtable_out["back_to_back_out"] == True])
    return nr_in, nr_out, back_to_back

def cleaning_numbers_period(df_bookingtable):
    """Make an table to use for the form of the cleaning company

    Args:
        df_bookingtable (_type_): _description_
    """ 
    date_to_check_from = st.sidebar.date_input("Date to check from_").strftime(
        "%Y-%m-%d"
    )  # .strptime("%Y-%m-%d")
    date_to_check_until = st.sidebar.date_input("Date to check until").strftime(
        "%Y-%m-%d"
    )  # .strptime("%Y-%m-%d")

    df_bookingtable_out = df_bookingtable[    (df_bookingtable["checkout_date"] >= date_to_check_from) & (df_bookingtable["checkout_date"] <= date_to_check_until)]
    
    df_bookingtable_out = df_bookingtable_out[
        ["checkout_date","acco_number", "guest_name", "acco_type","back_to_back_out"]
    ]

    # Define the conditions and corresponding values for the 'to_clean' column
    conditions = [
        df_bookingtable_out['acco_type'].isin(['BALI', 'WAIKIKI']),
        df_bookingtable_out['acco_type'].isin(['SERENGETI L', 'SERENGETI XL', 'KALAHARI1', 'KALAHARI2']),
        df_bookingtable_out['acco_type'].isin(['SAHARA'])
    ]
    values = ['MH', 'glamping', 'tent']

    # Apply the conditions and set the values in the 'to_clean' column
    df_bookingtable_out['cat_multiservice'] = np.select(conditions, values)
    
    if len(df_bookingtable_out) > 0:
        with st.expander("Total bookingtable"):
            st.table(df_bookingtable_out)

    pivot_table = df_bookingtable_out.pivot_table(index='checkout_date', columns='cat_multiservice', aggfunc='size').fillna(0)
    st.table(pivot_table)
    st.subheader(
        f"Total  = {len(df_bookingtable_out)}"
    )
   
    return 

def choose_start_end_date():

    import datetime
    date_to_check_from = st.sidebar.date_input("Date to check from")
    date_to_check_until = st.sidebar.date_input(
        "Date to check until", datetime.date(2023, 12, 31)
    )

    start_date = dt.datetime.combine(date_to_check_from, dt.datetime.min.time())
    end_date = dt.datetime.combine(date_to_check_until, dt.datetime.max.time())
    if start_date > end_date:
        st.error("Enddate cannot be before start date")
        st.stop()
    return start_date, end_date

def most_people_of_a_language(booking_df, language):
        
    # Assuming your booking table is stored in a DataFrame called 'booking_df'
    filtered_df = booking_df[booking_df['language'] == language]
    grouped = filtered_df.groupby('checkin_date')
    total_danish = grouped.size()
    total_checkins = booking_df.groupby('checkin_date').size()

    # Create a new DataFrame with the results
    result_df = pd.DataFrame({
        'Number of People': total_danish,
        'Total Check-ins': total_checkins
    })

    # Calculate relative percentage
    result_df['Relative Percentage'] = round(result_df['Number of People'] / result_df['Total Check-ins']*100,1)
    result_df = result_df.sort_values('Relative Percentage', ascending=False).fillna(0).reset_index()
    result_df['year'] = result_df['checkin_date'].str[:4]
    print (result_df.dtypes)
    print (result_df)
    

    # Display the results in Streamlit
    with st.expander("Dataframe"):
        st.write(result_df)

    result_df['date'] = pd.to_datetime(result_df['checkin_date'])
    result_df['month_day'] = result_df['date'].dt.strftime('%m-%d')  # Extract day and month in 'dd-mm' format
    result_df['month_day'] = pd.to_datetime(result_df['month_day'], format='%m-%d')  # Convert 'mm-dd' to date format
    #result_df['month_day'] = pd.to_datetime(result_df['month_day'], format='%m-%d').dt.strftime('%m-%d')  # Convert to 'mm-dd' format

    # Pivot the DataFrame to display the percentages in a table format
    for field in ['Number of People','Relative Percentage']:
    # Pivot the DataFrame to display the percentages in a table format
       
        pivot_df = result_df.pivot_table(values=field, index='month_day', columns='year').fillna(0).reset_index()

        # Display the pivot table in Streamlit
        st.subheader(f"{field} of People from {language}")

        
        # Define colors for each year
        colors = ['blue', 'red', 'green', 'orange']  # Add more colors if needed

        #Create traces for each year
        data = []
        for i,year in enumerate(pivot_df.columns[1:]):
            data.append(go.Scatter(x=pivot_df['month_day'], y=pivot_df[year], line=dict(color=colors[i % len(colors)]), name=str(year)))

        # Create the layout for the line graph
        layout = go.Layout(
            title=f"{field} People from {language} Over Time, based on checkin date",
            xaxis=dict(title="Check-in Date"),
            yaxis=dict(title=f"{field}"))
        

        # Create the figure
        fig = go.Figure(data=data, layout=layout)

        # Display the line graph in Streamlit
        st.plotly_chart(fig)



    def color_year(val):
        if val == "2019":
            color = 'red'
        elif val == "2021":
            color = 'yellow'
        elif val == "2022":
            color ='green'
        elif val == "2023":
            color = 'blue'    
        else:
            color = 'black'
       
        return f'background-color: {color}'

   

    # #st.table(result_df).style.applymap(color_year, subset=['year'])
    # #apply the function to the 'year' subset of the DataFrame
    # result_df_styled = result_df.style.apply(lambda x: color_year(result_df['year']), subset=['year'])
    # st.table(result_df_styled)

def show_bookingtable_period(df):
    """Show and save a bookingtable for a certain period
    
    """
    
    df = df.copy()
    df["checkin_date"] = pd.to_datetime(df["checkin_date"])
    df["checkout_date"] = pd.to_datetime(df["checkout_date"])
    df_to_show = df.copy()

    start_date, end_date =choose_start_end_date()
    df_to_show = df_to_show[(df_to_show["checkin_date"] >= start_date) & (df_to_show["checkout_date"] <= end_date)]
    st.write(df_to_show)

    st.write(f"Aantal : {len(df_to_show)}")
    
    # Convert the date columns to datetime type
    df["checkin_date"] = pd.to_datetime(df["checkin_date"])
    df["checkout_date"] = pd.to_datetime(df["checkout_date"])

   
    
    # Create a new DataFrame to store the summary table
    summary_df = pd.DataFrame(columns=["Date", "Number of Check-ins", "Number of Check-outs", "Number of Back to backs"])

    # Extract unique dates from the bookings
    dates = pd.date_range(start=df["checkin_date"].min(), end=df["checkout_date"].max(), freq="D")

    # Iterate over the dates and count the check-ins and check-outs
    for date in dates:
        checkin_count = len(df[df["checkin_date"] == date])
        checkout_count = len(df[df["checkout_date"] == date])
        back_to_back_in_count = len(df[(df["checkin_date"] == date) & (df["back_to_back_in"] == True)])
                       
        new_row = pd.DataFrame({"Date": [date], "Number of Check-ins": [checkin_count], "Number of Check-outs": [checkout_count], "Number of Back to backs":[back_to_back_in_count]})

        # Concatenate the new DataFrame with the existing summary_df
        summary_df = pd.concat([summary_df, new_row], ignore_index=True)

    summary_df = summary_df[
        (summary_df["Date"] <= end_date)
        & (summary_df["Date"] >= start_date)
    ]
    st.write(summary_df)

    for x_ in ["Number of Check-ins", "Number of Check-outs"]:
        fig = px.histogram(summary_df, x=x_, title=x_)
        st.plotly_chart(fig, use_container_width=True)
        afkapgrens = 20
        df_selection = summary_df[summary_df[x_] >=afkapgrens]
        st.write(f"Aantal {x_} groter of gelijk dan {afkapgrens} = {len(df_selection)}")

    for y_ in ["Number of Check-ins", "Number of Check-outs"]:
        fig = px.bar(summary_df, x="Date", y=y_, title=y_)
        st.plotly_chart(fig, use_container_width=True)

    return

def select_period(df_mutation, df_bookingtable):
  
    import datetime

    date_to_check_from = st.sidebar.date_input("Date to check from")
    date_to_check_until = st.sidebar.date_input(
        "Date to check until", datetime.date(2023, 12, 31)
    )

    start_date = dt.datetime.combine(date_to_check_from, dt.datetime.min.time())
    end_date = dt.datetime.combine(date_to_check_until, dt.datetime.max.time())
    if start_date > end_date:
        st.error("Enddate cannot be before start date")
        st.stop()

    # convert date columns to datetime format
    df_bookingtable["checkin_date"] = pd.to_datetime(df_bookingtable["checkin_date"])
    df_bookingtable["checkout_date"] = pd.to_datetime(df_bookingtable["checkout_date"])

    df_bookingtable = df_bookingtable[
        (df_bookingtable["checkin_date"] <= end_date)
        & (df_bookingtable["checkin_date"] >= start_date)
    ]
    
    df_mutation = df_mutation[
                (df_mutation["date"] <= end_date )
                  & (df_mutation["date"] >= start_date)
    ]
    return df_mutation, df_bookingtable

def make_checkin_outlist(df_bookingtable):
    date_to_check_from = st.sidebar.date_input("Date to check from_").strftime(
        "%Y-%m-%d"
    )  # .strptime("%Y-%m-%d")
    date_to_check_until = st.sidebar.date_input("Date to check until").strftime(
        "%Y-%m-%d"
    )  # .strptime("%Y-%m-%d")

    start_date = dt.datetime.strptime(date_to_check_from, "%Y-%m-%d").date()
    end_date = dt.datetime.strptime(date_to_check_until, "%Y-%m-%d").date()
    if start_date > end_date:
        st.error("Enddate cannot be before startdate")
        st.stop()
    current_datetime = dt.datetime.now()
    current_datetime_str = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

    st.header(f"Versie: {current_datetime_str}")
    # Iterate over each date between start and end dates
    current_date = start_date
    nr_in_totaal, nr_out_totaal,nr_b2b_totaal = 0, 0,0
    while current_date <= end_date:
        current_date_str = current_date.strftime("%Y-%m-%d")
        current_date_day_of_week = current_date.strftime("%a")

        #year_ = current_date.year
        #df_bookingtable = make_booking_table(year_)
        nr_in, nr_out, nr_b2b = checkin_outlist(
            df_bookingtable, current_date_str, current_date_day_of_week
        )

        # Move to the next date
        current_date += dt.timedelta(days=1)
        nr_in_totaal += nr_in
        nr_out_totaal += nr_out
        nr_b2b_totaal += nr_b2b
    st.header(f"Version: {current_datetime_str}")
    st.write(f"Period : {date_to_check_from} - {date_to_check_until}")
    st.write(f"Number of check-in = {nr_in_totaal} / Number of check-out = {nr_out_totaal} / Number of back to back: {nr_b2b_totaal}")

# Function to calculate Levenshtein distance between two strings
def levenshtein_distance(s, t):
    m, n = len(s), len(t)
    d = [[0] * (n + 1) for _ in range(m + 1)]

    for i in range(m + 1):
        d[i][0] = i
    for j in range(n + 1):
        d[0][j] = j

    for j in range(1, n + 1):
        for i in range(1, m + 1):
            if s[i - 1] == t[j - 1]:
                cost = 0
            else:
                cost = 1
            d[i][j] = min(d[i - 1][j] + 1, d[i][j - 1] + 1, d[i - 1][j - 1] + cost)

    return d[m][n]

# Function to compare last names and return a similarity score
def compare_last_names(row):
    #  ["Last Name","guest_name_booking", "Achternaam"],
    last_name_1 = row['Last Name']
    last_name_2 = row['guest_name_booking']
    distance = levenshtein_distance(last_name_1, last_name_2)
    max_length = max(len(last_name_1), len(last_name_2))
    similarity = 1 - (distance / max_length)
    return similarity

def find_unequal_rows(df, columm_maxxton, column_csv, name_test):

    """Find unequal rows/bookings. Helper function for compare_files()
    """
    st.subheader(f"Test to compare {name_test}")

    # in some rare cases there is a space after KALAHARI2 and KALAHARI1
    columns_to_strip = [columm_maxxton, column_csv]
    for column in columns_to_strip:
        if df[column].dtype == 'object':
            df[column] = df[column].str.strip()
            df[column] = df[column].str.upper()
    unequal_rows = df[df[columm_maxxton] != df[column_csv]]
    if name_test == "Land van herkomst":
        #unequal_rows = unequal_rows[~unequal_rows['guest_name'].str.startswith('miraculous')]
        unequal_rows = unequal_rows[unequal_rows['First Name'] != 'Miraculous']

    if name_test == "Land van herkomst (obv distributiekanaal)" or name_test == "Land van herkomst (obv country)":
        known_exceptions = [20945, 25927, 31007, 54432, 70038, 64860, 47840, 73200]
        unequal_rows = unequal_rows[~unequal_rows['Reservation Number'].isin(known_exceptions)]

    if name_test == "Accomodation type":
        unequal_rows["acco_type"] = unequal_rows["acco_type"].str.replace(' ', '_')
        unequal_rows['length_column_Accommodation Type'] = unequal_rows['Accommodation Type'].apply(lambda x: len(str(x)))
        unequal_rows['length_column_acco_type'] = unequal_rows["acco_type"].apply(lambda x: len(str(x)))

    if len(unequal_rows) == 0:
        st.write(f":white_check_mark: All {name_test} are the same")
    else:
        st.write(f":heavy_exclamation_mark: There are some unequal rows {name_test}")
        st.write(unequal_rows)
        with st.expander("Only the colums"):
            unequal_rows_x = unequal_rows[[columm_maxxton, column_csv]]
            st.write(unequal_rows_x)
        st.write(f"Number: {len(unequal_rows)}")

def compare_files(data_csv, data_maxxton):

    """Compare the maxxton file with the data in Maxxton.
    """    
   
    # SUPRESS WARNINGES
    #  SettingWithCopyWarning: 
    # A value is trying to be set on a copy of a slice from a DataFrame.
    # Try using .loc[row_indexer,col_indexer] = value instead

    # See the caveats in the documentation:
    # https://pandas.pydata.org/pandas-docs/stable/user_guide/indexing.html#returning-a-view-versus-a-copy
    data_maxxton = data_maxxton.copy()
    data_csv = data_csv.copy()

    # Filter the DataFrame to remove rows where accommodation type starts with 'pitch'
    data_maxxton = data_maxxton[~data_maxxton["Accommodation Type"].str.startswith("Pitch")]
    data_maxxton["Reservation Number"] = (
        data_maxxton["Reservation Number"].astype(str).str[3:8].astype(int)
    )
    data_maxxton["Arrival Date"] = pd.to_datetime(
        data_maxxton["Arrival Date"], format="%d/%m/%Y"
    )
    data_maxxton["Departure Date"] = pd.to_datetime(
        data_maxxton["Departure Date"], format="%d/%m/%Y"
    )
    try:
        data_maxxton["Booking country"] = data_maxxton["Distribution Channel"].str[-2:]
    except Exception:
        print ("There is no column for distribution Channel in file")
    # Mapping of original values to replacement values
    value_map_acc = {
        "Safari tent Serengeti XL Glamping Soleil": "SERENGETI XL",
        "Safari tent Kalahari Soleil [5 pers. (5 adults) 32m²": "KALAHARI1",
        "Safari tent Serengeti Glamping Soleil": "SERENGETI L",
        "Mobile home Waikiki Soleil": "WAIKIKI",
        "Safari tent Kalahari Soleil [5 pers. (5 adults) 25m²": "KALAHARI2",
        "Mobile home Bali Soleil": "BALI",
        "Bungalow tent Navajo Soleil": "SAHARA",
    }

    # Replace values in the 'Property' column
    for original_value, replacement_value in value_map_acc.items():
        data_maxxton.loc[
            data_maxxton["Accommodation Type"].str.startswith(original_value), "Accommodation Type",] = replacement_value
    try:
        value_map_country = {
            "Belgium":"be",
            "Switserland":"ch",
            "Germany":"de",
            "Denmark":"dk",
            "France":"fr",
            "Great Britain":"en",
            "Luxembourg":"lx",
            "Netherlands":"nl",
            "Philippines (the)":"ph"}
        for original_value, replacement_value in value_map_country.items():
            data_maxxton.loc[
                data_maxxton["Country"].str.startswith(original_value), "Country",] = replacement_value.upper()
    except Exception:
        print ("There is no column for Country / Distribution Channel in file")
        

    #data_csv = make_bookingtable_period(data_csv)  # pd.read_csv(bookingtable)
    data_csv["booking_number"] = data_csv["booking_number"].astype(int)
    data_csv["checkin_date"] = pd.to_datetime(
        data_csv["checkin_date"], format="%Y-%m-%d"
    )
    data_csv["checkout_date"] = pd.to_datetime(
        data_csv["checkout_date"], format="%Y-%m-%d"
    )
    data_csv = data_csv[data_csv["booking_number"] != 0]
    # data_csv["language"] = data_csv["language"].str.upper()
    
    data_csv["language"] = data_csv.loc[:,"language"].str.upper()
    df = pd.merge(
        data_maxxton,
        data_csv,
        how="inner",
        left_on="Reservation Number",
        right_on="booking_number",
    )
    df = df[df["First Name"] != "Miraculous"]

    to_compare = [["Accommodation Type","acco_type", "Accomodation type"],
        ["Arrival Date","checkin_date", "Check in date"],
       
        ["Departure Date", "checkout_date", "check Out date"],
        # ["Last Name","guest_name_booking", "Achternaam"],
        ["Booking country", "language", "Land van herkomst (obv distributiekanaal)"],
        ["Country", "language", "Land van herkomst (obv country)" ]]

    for c in to_compare:
        try:
            find_unequal_rows(df, c[0], c[1], c[2])
        except Exception:
            st.write(f":question: Cannot compare {c[2]} - probably column missing in one of the files")
        
    
    st.subheader("Last names - levenshtein_distance")
    # Create a new column 'similarity_score' in the DataFrame to store the similarity scores
    df['similarity_score'] = df.apply(compare_last_names, axis=1)

    # Set a threshold for similarity score (e.g., 0.8)
    threshold = 0.3

    # Filter the DataFrame to get rows where the similarity score is below the threshold
    mismatched_rows = df[df['similarity_score'] < threshold]
    matched_rows = df[df['similarity_score'] >= threshold]
    if len(mismatched_rows) == 0:
        st.write(":white_check_mark: All last names are almost the same")
    else:
    # Print the mismatched rows
        mismatched_rows = mismatched_rows.sort_values("similarity_score", ascending=True)
        mismatched_rows_x = mismatched_rows[["Reservation Number", "Last Name","guest_name_booking", 'similarity_score']]
        mismatched_rows_x = mismatched_rows_x.sort_values("similarity_score", ascending=True)
        
        st.write(mismatched_rows_x)

        with st.expander("All colums"):
            st.write(mismatched_rows)
            
        st.write(f"Number: {len(mismatched_rows)}")

        with st.expander("The matched rows"):
            matched_rows_x = matched_rows[["Last Name","guest_name_booking"]]
            st.write(matched_rows_x)
            st.write(f"Number: {len(matched_rows)}")
    # Perform an anti-join to get rows where there is no common element
    anti_join = data_maxxton[
        ~data_maxxton["Reservation Number"].isin(data_csv["booking_number"])
    ]
    if len(anti_join) == 0:
        st.subheader("Maxxton vs Excel")
        st.write(":white_check_mark: All reservations in Maxxton are in registred in Excel")

    else:
        st.subheader(":question: Reservations in Maxton but not in Excel (forgotten to put in?)")

        st.write(anti_join)
        st.write(f"Number: {len(anti_join)}")

    # Perform an anti-join to get rows where there is no common element
    # anti_join2 = data_maxxton[~data_maxxton['Reservation Number'].isin(data_csv['booking_number'])]
    anti_join2 = data_csv[
        ~data_csv["booking_number"].isin(data_maxxton["Reservation Number"])
    ]
    if len(anti_join2) == 0:
        st.subheader("Excel vs Maxxton")
        st.write(":white_check_mark: All reservations in Excel are in Maxxton")
    else:
        st.subheader(":question: Reservations in the Excel file but not in Maxxton (cancelled or not paid?)")

        st.write(anti_join2)
        st.write(f"Number: {len(anti_join2)}")

        # Assuming your DataFrame is called 'df'
    
        
    # Assuming your DataFrame is called 'df'
    desired_columns = ['acco_number', 'guest_name', 'checkin_date', 'checkout_date', 'Customer Due Amount']
    # Get the remaining columns (excluding the desired columns)
    remaining_columns = [col for col in df.columns if col not in desired_columns]
    # Rearrange the columns by concatenating the desired columns followed by the remaining columns
    new_columns = desired_columns + remaining_columns
    # Reorder the DataFrame based on the new column order
    df = df[new_columns]

    df_with_xxx =  df[df['guest_name'].str.contains('xxx')]
    if len(df_with_xxx) == 0:
        st.subheader("Guest with open bill according to Excel")
        st.write(":white_check_mark: All reservations are paid")
    else:
        st.subheader(":question: Reservations marked as not paid in Excel")
        st.write(df_with_xxx.sort_values('Customer Due Amount'))
        st.write(f"Number: {len(df_with_xxx)}")

    df_with_pos_cda = df[df['Customer Due Amount'] > 0]
    if len(df_with_pos_cda) == 0:
        st.subheader("Guest with Customer Due Amount in Maxxton")
        st.write(":white_check_mark: All guests in Maxxton have paid")
    else:
        st.subheader(":question: Guests with Customer Due Amount in Maxxton - still have to pay")
        st.write(df_with_pos_cda.sort_values('Customer Due Amount'))
        st.write(f"Number: {len(df_with_pos_cda)}")

    df_with_neg_cda = df[df['Customer Due Amount'] < 0]
    if len(df_with_neg_cda) == 0:
        st.subheader("Guest having negative Customer Due Amount in Maxxton")
        st.write(":white_check_mark: No guests with negative CDA")
    else:
        st.subheader(":question: Guests having negative Customer Due Amount in Maxxton")
        st.write(df_with_neg_cda.sort_values('Customer Due Amount'))
        st.write(f"Number: {len(df_with_neg_cda)}")
        
def add_on_list(data_csv):
    """
    Generate add-ons summaries and visualizations based on the given CSV data.

    Args:
        data_csv (pandas.DataFrame): A DataFrame containing the data.

    Returns:
        None
    """

    #data_csv = make_bookingtable_period()  # pd.read_csv(bookingtable)
    data_csv = data_csv[["checkin_date", "acco_number", "guest_name", "booking_number", "sng_linnen", "dbl_linnen", "kst", "bb"]]
    
    # Create a boolean mask indicating rows with non-zero values in any of the specified columns
    mask = (data_csv['sng_linnen'] != 0) | (data_csv['dbl_linnen'] != 0) | (data_csv['kst'] != 0) | (data_csv['bb'] != 0)

    # Apply the mask to filter the DataFrame
    filtered_data = data_csv[mask]

    # Sort the filtered DataFrame by 'checkin_date'
    filtered_data = filtered_data.sort_values(by='checkin_date')
    st.subheader("Add-ons per booking")
    st.write(filtered_data)

    # Group by 'checkin_date' and calculate the sum of the columns
    sum_table = filtered_data.groupby('checkin_date')[['sng_linnen', 'dbl_linnen', 'kst', 'bb']].sum()
    
    # Display the new table with the summed values
    st.subheader("Add-ons per day")
    st.write(sum_table)

    # Set 'checkin_date' column as the index
    filtered_data.set_index('checkin_date', inplace=True)
    
    # Resample the data by week, starting from Monday (M) and ending on Sunday (W-SUN)
    weekly_table = filtered_data.resample('W-MON')['sng_linnen', 'dbl_linnen', 'kst', 'bb'].sum()
    
    # Display the new table with summed values per week
    st.subheader("Add-ons per week")
    #st.write(weekly_table)

    # Calculate the sum of all rows in the weekly table
    total_sum = weekly_table[['sng_linnen', 'dbl_linnen', 'kst', 'bb']].sum(axis=0)

    # Create a new DataFrame with the total sum as an extra row
    total_row = pd.DataFrame([total_sum], columns=weekly_table.columns, index=['Total'])
    table_with_total = pd.concat([weekly_table, total_row])
    
    # Display the table with the total sum as an extra row
    st.write(table_with_total) 
    #Serialization of dataframe to Arrow table was unsuccessful due to: 
    # ("object of type <class 'str'> cannot be converted to int", 
    # 'Conversion failed for column None with type object').
    #  Applying automatic fixes for column types to make the dataframe Arrow-compatible.

    
def make_and_show_length_of_stay(df, df_bookingtable,start_month,end_month):

    def show_verblijfsduur_per_month_per_jaar(df_bookingtable):
        """_summary_

        Args:
            df_bookingtable (_type_): _description_
        """

      
            # Step 1: Filter relevant columns
        df_filtered = df_bookingtable[['acco_type', 'checkin_date', 'month','year', 'number_of_days']].copy()




        df_filtered["one"] = 1
        grouped_df = df_filtered.groupby(['month', 'year']).agg({'number_of_days': 'sum', 'one': 'sum'}).reset_index()
        # Step 2: Convert 'checkin_date' to datetime
        #grouped_df['checkin_date'] = pd.to_datetime(grouped_df['checkin_date'])

        grouped_df["number_of_days_avg"] = round((grouped_df["number_of_days"] / grouped_df["one"]),1)
        
        # Step 4: Create pivot table
        pivot_table = grouped_df.pivot_table(index='month', columns='year', values='number_of_days_avg', fill_value=0)
        # Step 5: Round the values in the pivot table
        pivot_table = pivot_table.round(1)  # Round to 2 decimal places
        
        st.subheader(f"Average length of stay")
        st.write(pivot_table)
   
    def show_frequency_table_duration_of_stay(df_bookingtable, month, normalized, cumulative, start_month,end_month):
        
        """_summary_

        Args:
            df_bookingtable (df): _description_
            month (int): _description_
            normalized (boolean): _description_
            cumulative(boolean)
            start_month(int): just for the title of the table
            end_month(int):just for the title of the table
        """        
        df_filtered = df_bookingtable[['acco_type', 'checkin_date', 'month','year', 'number_of_days']].copy()


      

        if month!=None:
            df_month_year = df_filtered[(df_filtered["month"] == month) ]
            title_month = month
        else:
            df_month_year = df_filtered
            if start_month == end_month:
                title_month = f"month: {start_month}"
            else:
                title_month = f"month: {start_month} until {end_month} (incl)"
       
        
        # Create a new DataFrame with the desired frequency table
        frequency_table = df_month_year.pivot_table(index='number_of_days', columns='year', aggfunc='size', fill_value=0)
        st.subheader(f"Frequency table - {title_month}")
        st.write(frequency_table)
        # Normalize the frequency table
        normalized_table = frequency_table.apply(lambda x: round((x / x.sum() * 100),1), axis=0)
        st.subheader(f"Frequency table - percentages - {title_month}")
        # Display the normalized table
        st.write(normalized_table)
        st.subheader(f"Frequency table - cummulative - {title_month}")
        # Compute the cumulative sum
        cumulative_table = normalized_table.cumsum()

        # Add a row with 100% at the bottom
        cumulative_table.loc['Total'] = 100

        # Display the cumulative table
        st.write(cumulative_table)

    def show_histogram_verblijfsduur(df_bookingtable, month, year, normalized, cumulative):
        """_summary_

        Args:
            df_bookingtable (df): _description_
            month (int): _description_
            year (str): _description_
            normalized (boolean): _description_
            cumulative(boolean)
        """        
        df_filtered = df_bookingtable[['acco_type', 'checkin_date', 'month','year', 'number_of_days']].copy()


      

        if month!=None:
            df_month_year = df_filtered[(df_filtered["month"] == month) & (df_filtered["year"] == year) ]
        else:
            df_month_year = df_filtered[(df_filtered["year"] == year) ]
        
        
        
        if normalized:
            fig = go.Figure(data=[go.Histogram(x=df_month_year['number_of_days'], histnorm='probability', cumulative_enabled=cumulative)])
        
            # Set the title and labels for the graph
            fig.update_layout(title=f'Histogram of Number of Days (Relative) in {month} / {year}',
                                    xaxis_title='Number of Days',
                                    yaxis_title='Probability')
           
        else: 
            fig = go.Figure(data=[go.Histogram(x=df_month_year['number_of_days'], cumulative_enabled=cumulative)])
            fig.update_layout(title=f'Histogram of Number of Days in {month} / {year}',
                            xaxis_title='Number of Days',
                            yaxis_title='Count')
     
        st.plotly_chart(fig)
   


    def show_verblijfsduur_aantal_boekingen_per_month(df_bookingtable):
        """_summary_

        Args:
            df_bookingtable (_type_): _description_
        """

        # Step 1: Filter relevant columns
        df_filtered = df_bookingtable[['acco_type', 'checkin_date', 'month','year', 'number_of_days']].copy()

        # Step 2: Convert 'checkin_date' to datetime
        df_filtered['checkin_date'] = pd.to_datetime(df_filtered['checkin_date'])

        # Iterate over unique years
        unique_years = df_filtered['year'].unique()
        for year in unique_years:
            # Step 3: Filter by year
            df_year = df_filtered[df_filtered['year'] == year]

            # Step 4: Create pivot table
            # TODO : adjust like show_verblijfsduur_per_month_per_jaar(df_bookingtable)
            pivot_table = df_year.pivot_table(index='acco_type', columns='month', values='number_of_days', aggfunc='mean', fill_value=0)
            # Step 5: Round the values in the pivot table
            pivot_table = pivot_table.round(1)  # Round to 2 decimal places

            # Step 4: Create pivot table
            pivot_table_aantal = df_year.pivot_table(index='acco_type', columns='month', values='number_of_days', aggfunc='count', fill_value=0)

            # Step 5: Do something with the pivot table for each year
            
            st.subheader(f"{year} - Gemidelde verblijfsduur")
            st.write(pivot_table)
            st.subheader(f"{year} - Aantal boekingen")
            st.write(pivot_table_aantal)

    def show_average_stay_per_accotype_per_month(df_bookingtable):
        """show_average_stay_per_accotype_per_month

        Args:
            df_bookingtable (_type_): _description_
        """        
        # Step 1: Filter relevant columns
        df_filtered = df_bookingtable[['acco_type', 'checkin_date', 'month','year', 'number_of_days']].copy()

        # Step 2: Convert 'checkin_date' to datetime
        df_filtered['checkin_date'] = pd.to_datetime(df_filtered['checkin_date'])

        # Step 3: Get unique years and months
        unique_years = df_filtered['year'].unique()
        unique_months = df_filtered['month'].unique()
        unique_months.sort()
        
        # Step 4: Iterate over acco_types
        unique_acco_types = df_filtered['acco_type'].unique()
        for acco_type in unique_acco_types:
            # Create an empty DataFrame to store the pivot table for each acco_type
            pivot_table = pd.DataFrame(index=unique_years, columns=unique_months)

            # Step 5: Iterate over years
            for year in unique_years:
                # Step 6: Filter by year and acco_type
                df_year_acco = df_filtered[(df_filtered['year'] == year) & (df_filtered['acco_type'] == acco_type)]

                # Step 7: Calculate average stay for each month
                for month in unique_months:
                    avg_stay = df_year_acco[df_year_acco['month'] == month]['number_of_days'].mean()
                    pivot_table.loc[year, month] = round(avg_stay, 2)
            st.subheader(f"Average stay {acco_type}")
            st.write(pivot_table)
 
    show_verblijfsduur_per_month_per_jaar(df_bookingtable)
    
    normalized, cumulative, month = True, True, None
    show_frequency_table_duration_of_stay(df_bookingtable, month,  normalized, cumulative,start_month,end_month)
    for normalized in [False,True]:
      
        for cumulative in [False,True]:
            show_histogram_verblijfsduur(df_bookingtable, month, "2021",normalized, cumulative)
            show_histogram_verblijfsduur(df_bookingtable, month, "2022",normalized, cumulative)
            show_histogram_verblijfsduur(df_bookingtable, month, "2023",normalized, cumulative)
    show_verblijfsduur_aantal_boekingen_per_month(df_bookingtable)
    show_average_stay_per_accotype_per_month(df_bookingtable)

def make_and_show_pivot_tables(df, df_bookingtable,start_month,end_month):
  
    def show_number_per_month_per_jaar(df_bookingtable):
        """_summary_

        Args:
            df_bookingtable (_type_): _description_
        """

      
            # Step 1: Filter relevant columns
        df_filtered = df_bookingtable[['acco_type', 'checkin_date', 'month','year', 'number_of_days']].copy()

        df_filtered["one"] = 1
        grouped_df = df_filtered.groupby(['month', 'year']).agg({'number_of_days': 'sum', 'one': 'sum'}).reset_index()
        grouped_df["number_of_days_avg"] = round((grouped_df["number_of_days"] / grouped_df["one"]),1)
        
        # Step 4: Create pivot table
        pivot_table_stays = grouped_df.pivot_table(index='month', columns='year', values='one', fill_value=0).round(1)
        st.subheader(f"Number of bookings")
        st.write(pivot_table_stays)

        pivot_table_nights = grouped_df.pivot_table(index='month', columns='year', values='number_of_days', fill_value=0).round(1)       
        st.subheader(f"Number of nights")
        st.write(pivot_table_nights)

        pivot_table_avg = grouped_df.pivot_table(index='month', columns='year', values='number_of_days_avg', fill_value=0).round(1)       
        st.subheader(f"Number of days avg")
        st.write(pivot_table_avg)
    
    def show_omzet_per_month_per_year(df):
        """
        Displays the omzet (revenue) per month (month) per year (year) as a pivot table.

        Args:
            df (pandas.DataFrame): The input DataFrame containing the data.

        Returns:
            None: The resulting pivot table is displayed using the st.write() function.

        Example:
            show_omzet_per_month_per_year(df)
        """
        st.subheader("Omzet per month per year")
        pivot_df = pd.pivot_table(df, values='omzet', index='month', columns='year', aggfunc='sum')
        # Display the resulting pivot table
        st.write(pivot_df)

    def show_occupation_per_month_per_year(df):
        """_summary_

        Args:
            df (_type_): _description_
        """    
        st.subheader("occupation per month per year")
        # Assuming your dataframe is named 'df'
        
        grouped_df = df.groupby(['year', 'month']).agg({'in_house': 'sum', 'aantal': 'sum'})
        grouped_df['occupation'] = round( grouped_df['in_house'] / grouped_df['aantal']*100,1)
        pivot_df = pd.pivot_table(grouped_df, values='occupation', index='month', columns='year').fillna(0)

        # Display the resulting pivot table
        st.write(pivot_df)

     
    
    def show_occupation_per_accotype_per_month(df):
        """_summary_

        Args:
            df (_type_): _description_
        """    
        # Assuming your dataframe is named 'df'
        for y in [2023]: #2019,2021,2022,
            #st.subheader(f"{y}")
            # Filter dataframe for the desired year
            filtered_df = df[df['year'] == str(y)]
            # st.write(filtered_df)
            # Group by month and acco_type, calculate occupation
            grouped_df = filtered_df.groupby(['month', 'acco_type']).agg({'in_house': 'sum', 'aantal': 'sum', "omzet": 'sum'})
            grouped_df['occupation'] =  round( grouped_df['in_house'] / grouped_df['aantal']*100,1)

            # Create pivot table
            pivot_df = pd.pivot_table(grouped_df, values='occupation', index='month', columns='acco_type')
            st.subheader(f"occupation  per accotype, per month in {y}")
            # Display the resulting pivot table
            st.write(pivot_df)

            # Create pivot table
            pivot_df_omzet = pd.pivot_table(grouped_df, values='omzet', index='month', columns='acco_type')
            st.subheader(f"Turn over in {y}")
            # Display the resulting pivot table
            st.write(pivot_df_omzet)
   
    show_number_per_month_per_jaar(df_bookingtable)
    show_omzet_per_month_per_year(df)
    show_occupation_per_month_per_year(df)

 
    show_occupation_per_accotype_per_month(df)

def show_checkin_info(df_mutation):
    
    def show_checkins_per_month_per_year(df_mutation):
        df_mutation["check_ins"] = df_mutation["new_arrival"]+df_mutation["back_to_back"]
        
        # Group the dataframe by month
        df_grouped = df_mutation.groupby([ 'year', 'month']).agg({'check_ins': 'sum'}).reset_index()

        # # Create the pivot table
        # pivot_table = pd.pivot_table(df_grouped, values='check_ins', index='month', columns='year', aggfunc='sum').fillna(0)

        #Create the pivot table
        pivot_table = pd.pivot_table(df_grouped, values='check_ins', index='month', columns='year', aggfunc='sum', margins=True, margins_name='Total').fillna(0)

        # Calculate the row averages
        # pivot_table['Average'] = pivot_table.mean(axis=1) 


        st.subheader("show_checkins_per_month")
        st.write(pivot_table)

    def show_avg_checkins_per_day_per_month_per_year(df_mutation):
        df_mutation["check_ins"] = df_mutation["new_arrival"] + df_mutation["back_to_back"]
        df_mutation["day_of_week"] = df_mutation["date"].dt.dayofweek

        df_grouped = df_mutation.groupby(['date', 'month', 'year_int', 'day_of_week']).agg({'check_ins': 'sum'}).reset_index()
        df_grouped["is_wednesday"] = np.where(df_grouped["year_int"] >= 2023, np.where(df_grouped["day_of_week"] == 2, 1, 0), 0)

        df_grouped = df_grouped.groupby(['year_int', 'month']).agg({'check_ins': 'sum', 'is_wednesday': 'sum'}).reset_index()

        df_grouped['days_in_month'] = df_grouped.apply(lambda row: calendar.monthrange(int(row['year_int']), row['month'])[1], axis=1)
        df_grouped['avg_check_ins_per_day'] = df_grouped['check_ins'] / (df_grouped['days_in_month'] - df_grouped['is_wednesday'])

        pivot_table = pd.pivot_table(df_grouped, values='avg_check_ins_per_day', index='month', columns='year_int', aggfunc='sum').fillna(0).round(1)
        st.subheader("show avg check-ins per day per month per year")
        st.write(pivot_table)

    def show_avg_checkins_per_weekday_per_month_per_year(df_mutation):
        df_mutation["check_ins"] = df_mutation["new_arrival"] + df_mutation["back_to_back"]
        df_mutation["day_of_week"] = df_mutation["date"].dt.dayofweek
        df_mutation["day_name"] = df_mutation["date"].dt.day_name()
        df_grouped = df_mutation.groupby(['date', 'month', 'year_int', 'day_of_week', 'day_name']).agg({'check_ins': 'sum'}).reset_index()
        df_grouped["one"] = 1
        df_grouped = df_grouped.groupby(['year_int', 'month', 'day_of_week', 'day_name']).agg({'check_ins': 'sum', 'one': 'sum'}).reset_index()
        df_grouped['avg_check_ins_per_weekday'] = df_grouped['check_ins'] / df_grouped['one'] 

        pivot_table = pd.pivot_table(df_grouped, values='avg_check_ins_per_weekday', index=['month', 'day_of_week', 'day_name'], columns='year_int', aggfunc='mean').fillna(0).round(1)
        st.subheader("Show average check-ins per weekday per month")
        st.write(pivot_table)


    def graph_number_of_checkins_per_day(df_mutation):
        st.subheader("Number of checkins per day, all years")
        df_mutation["check_ins"] = df_mutation["new_arrival"]+df_mutation["back_to_back"]

        df_mutation_grouped_by_date_year = df_mutation.groupby(["year","date", "day_month"])[["check_ins"]].sum().sort_values(by="date").reset_index()    
        df_mutation_grouped_by_date_year["day_month_dt"] = pd.to_datetime(df_mutation_grouped_by_date_year["day_month"], format="%d-%m")
        pivot_table = df_mutation_grouped_by_date_year.pivot_table(index="day_month_dt", columns="year", values="check_ins")

        figy = go.Figure()
        # Define colors for each year
        colors = ['red', 'green', 'orange']  # Add more colors if needed

        #Create traces for each year
        data = []
        for i,year in enumerate(pivot_table.columns[1:]): #not 2019
            data.append(go.Scatter(x=pivot_table.index, y=pivot_table[year], line=dict(color=colors[i % len(colors)]), name=str(year)))
            
        #figy.update_layout(xaxis_tickformat="%d-%b")  # Display only the day and month on x-axis
        layout = go.Layout(
            title=f"Number of checkins per checkin date",
            xaxis=dict(title="Check-in Date"),
            yaxis=dict(title=f"Nuber of checkins"))
        
        
        # Create the figure
        figy = go.Figure(data=data, layout=layout)
        figy.update_layout(xaxis_tickformat="%d-%b")
        st.plotly_chart(figy, use_container_width=True)

        figz = go.Figure()
        #Create traces for each year
        for i,year in enumerate(pivot_table.columns[1:]): #not 2019
            figz.add_trace(go.Bar(x=pivot_table.index, y=pivot_table[year], name=str(year)))

        figz.update_layout(xaxis_tickformat="%d-%b")  # Display only the day and month on x-axis
        st.plotly_chart(figz, use_container_width=True)

    def show_busy_days_per_month_per_year(df_mutation, busy_factor):
        
        df_mutation["check_ins"] = df_mutation["new_arrival"]+df_mutation["back_to_back"]
      
        # Group the dataframe by month
        df_grouped = df_mutation.groupby('date').agg({'check_ins': 'sum'})
        df_grouped = df_grouped[df_grouped["check_ins"] >= busy_factor].reset_index()
        df_grouped['month'] = df_grouped['date'].dt.month
        df_grouped['year'] = df_grouped['date'].dt.year
        df_grouped['one']  = 1
        pivot_table = pd.pivot_table(df_grouped, values='one', index='month', columns='year', aggfunc='sum').fillna(0)
        months_range = range(3, 11)  # Assuming data spans all 12 months of the year
        pivot_table = pivot_table.reindex(months_range, fill_value=0)
        # Create the pivot table
        st.subheader(f"Number of days with more than {busy_factor} checkins per day")
        st.write(pivot_table)

    show_checkins_per_month_per_year(df_mutation)
    show_avg_checkins_per_day_per_month_per_year(df_mutation)
    show_avg_checkins_per_weekday_per_month_per_year(df_mutation)
    graph_number_of_checkins_per_day(df_mutation)
    for busyfactor in [10,15,20,25,30]:
        show_busy_days_per_month_per_year(df_mutation,busyfactor)


@st.cache_data()
def get_data_local():
    """Get the data if the script is run locally

    Returns:
        df: the various dataframes
    """    
    excel_file_2023 = r"C:\Users\rcxsm\Downloads\bezetting2023a.xlsm"
    maxxton_file = r"C:\Users\rcxsm\Downloads\ReservationDetails.xlsx"
       
    s1 = int(time.time())
    wb_2023 = load_workbook(excel_file_2023, data_only=True) 
    df_maxxton = pd.read_excel(maxxton_file)   
    
    df_mutation= make_mutation_df(wb_2023)
    df_bookingtable = make_booking_table(wb_2023)
    s2 = int(time.time())
    s2x = s2 - s1
    print(" ")  # to compensate the  sys.stdout.flush()
    print(f"Uploading and making df_mutation and df_booking  took {str(s2x)} seconds ....)")
    return df_mutation, df_bookingtable, df_maxxton
    

#@st.cache_data()
def upload_files():

    """ Uploads the files

            Streamlit manages your uploaded files for you. Files are stored in memory (i.e. RAM, not disk), 
            and they get deleted immediately as soon as they’re not needed anymore.

            This means we remove a file from memory when:

            The user uploads another file, replacing the original one
            The user clears the file uploader
            The user closes the browser tab where they uploaded the file
            https://discuss.streamlit.io/t/where-does-the-data-go-when-using-file-uploader-when-does-it-get-deleted/8269/1

    Returns:
        wb_2023: the workbook
        df_maxxton : the df with the info from the Maxxton output

    """    
    excel_file_2023 = st.sidebar.file_uploader("Choose the Henriette file", type='xlsm')
    maxxton_file = st.sidebar.file_uploader("Choose the Maxxton file", type='xlsx')
    if (excel_file_2023 is not None) and (maxxton_file is not None):
        s1 = int(time.time())
        wb_2023 = load_workbook(excel_file_2023, data_only=True) 
        df_maxxton = pd.read_excel(maxxton_file)   
        s2 = int(time.time())
        s2x = s2 - s1
       
        print(f"Uploading  took {str(s2x)} seconds ....)")
        return wb_2023, df_maxxton
    else:
        st.error("Please upload the files")
        st.stop()

def get_data(wb_2023, df_maxxton):
        s1 = int(time.time())
        df_mutation= make_mutation_df(wb_2023)

        df_bookingtable = make_booking_table(wb_2023)
        s2 = int(time.time())
        s2x = s2 - s1
        print(" ")  # to compensate the  sys.stdout.flush()
        print(f"Making df_mutation and df_booking  took {str(s2x)} seconds ....)")
        return df_mutation, df_bookingtable, df_maxxton

@st.cache_data()
def make_cache_data(df_mutation, df_bookingtable, df_maxxton):

    """Making a cache with the various dataframes. 

    Caching get_dat() gives this error : 
        Cannot hash argument 'wb_2023' (of type `openpyxl.workbook.workbook.Workbook`) in 'get_data'.

        To address this, you can tell Streamlit not to hash this argument by adding a
        leading underscore to the argument's name in the function signature:

        ```
        @st.cache_data
        def get_data(_wb_2023, ...):
            ...
        ```

    Returns:
        _type_: the same dataframes
    """    
    return df_mutation, df_bookingtable, df_maxxton

def filter_months(start_month, end_month, df_):
    df_["month"] = df_["month"].astype(int)
    df_ = df_[(df_["month"] >= start_month) & (df_["month"] <= end_month)]
    return df_

def main():       
    test = True  # To test or not to test (to see if the fillcolors in the sheet are right.)
    # https://github.com/onedrive/onedrive-sdk-python
    # https://github.com/vgrem/Office365-REST-Python-Client#Working-with-OneDrive-API
    upload = True
    if platform.processor() != "" and upload == False:
        df_mutation, df_bookingtable, df_maxxton = get_data_local()
    else:

        wb_2023, df_maxxton = upload_files()
        
        df_mutation, df_bookingtable, df_maxxton = get_data(wb_2023, df_maxxton)
        df_mutation, df_bookingtable, df_maxxton = make_cache_data(df_mutation, df_bookingtable, df_maxxton)
   
    ix = 1 if platform.processor() != "" else  0
    
    keuze = st.sidebar.selectbox("Compare Files / Complete menu", ["Compare Files", "Complete menu"], index=ix)

    if keuze == "Compare Files":
        what_to_do =  "Compare files"
    else:
        what_to_do = st.sidebar.radio(
            "What to do",
            (
                "Hello",
                "--- OPERATIONS ---",
                "Checkin/out list",
                "Bookingtable",
                "Add-on list",
                "Cleaning numbers month",
                "---- INTELLIGENCE ---",
                "Analyse",
                "pivot tables",
                "analyse of number of checkins",
                "info number of days",
                "length of stay",
                "occupation graph",
                
                "show info from bookingtable",

                
            

                "dekenanalyse",
                "babypackanalyse",

                "most people from xx",
                
                "---- UTILITIES ---",
                "Find color",
                "Compare files",
                "Check Excel File", 
                "Clear cache",
                
            ),
            0,
        )

  
    
    # if what_to_do == "Analyse" or what_to_do == "Check Excel File":
    #     start_month, end_month = 3, 11

    #     (start_month, end_month) = st.sidebar.slider(
    #         "Months (from/until (incl.))", 1, 12, (3, 10)
    #     )

    #     if start_month > end_month:
    #         st.warning("Make sure that the end month is not before the start month")
    #         st.stop()

    selection_list_accos = [
        "WAIKIKI",
        "BALI",
        "KALAHARI1",
        "KALAHARI2",
        "SERENGETI XL",
        "SERENGETI L",
        "SAHARA",
    ]
    #     selection_list_accos = st.sidebar.multiselect(
    #         "Welke acco's", selection_list_accos_, selection_list_accos_
    #     )
    years = [2019, 2021, 2022, 2023]
    # years = st.sidebar.multiselect("Jaren", year_, [2021, 2022, 2023])
    (start_month,end_month) = st.sidebar.slider("Months (from/until (incl.))", 1, 12, (3,10))
   
    df_mutation = filter_months(start_month, end_month, df_mutation)
    
    df_bookingtable = filter_months(start_month, end_month, df_bookingtable)
   
    print (f"--- ({what_to_do}) ---")
    if what_to_do.startswith("---"):
        pass
    else:
        st.header(what_to_do)
    if what_to_do in ["Add-on list","Compare files"]:
        df_mutation_period, df_bookingtable_period = select_period(df_mutation, df_bookingtable)

    if what_to_do == "Hello" or what_to_do.startswith("---"):
        st.write("Hello")
        st.stop()
   
    # OPERATIONS
    elif what_to_do == "Checkin/out list":    
        make_checkin_outlist(df_bookingtable)
    elif what_to_do == "Bookingtable":
        show_bookingtable_period(df_bookingtable)    
    elif what_to_do == "Add-on list":
        add_on_list(df_bookingtable_period)

    elif what_to_do == "Cleaning numbers month":
        cleaning_numbers_period(df_bookingtable)

    # BUSINIESS INTELLIGENCE
    elif what_to_do == "Analyse":
        for y in years:
            st.subheader(y)
          
            df_mutations_year = df_mutation[df_mutation["year"] == str(y)]
            generate_businessinfo(df_mutations_year)
    elif what_to_do == "pivot tables":
        make_and_show_pivot_tables(df_mutation, df_bookingtable, start_month,end_month)
    elif what_to_do == "analyse of number of checkins":
        show_checkin_info(df_mutation)
    elif what_to_do == "length of stay":  
        make_and_show_length_of_stay(df_mutation, df_bookingtable,start_month,end_month)

    elif what_to_do == "occupation graph":
        make_occupuation_graph(df_mutation)
    
    
    elif what_to_do == "show info from bookingtable":
        for y in years:
            df_bookingtable_year =  df_bookingtable[df_bookingtable["year_int"] == y]
            show_info_from_bookingtable(df_bookingtable_year, y)
    elif what_to_do == "info number of days":
        for y in years:
            df_bookingtable_year =  df_bookingtable[df_bookingtable["year_int"] == y]
            show_info_number_of_days(df_bookingtable_year,y)
    elif what_to_do == "most people from xx":
        st.subheader("All time")
        for l in ["dk","de", "en"]:
            most_people_of_a_language(df_bookingtable, l)
      
        # for y in years:
        #     st.subheader(f"{y}")
        #     df_bookingtable_year =  df_bookingtable[df_bookingtable["year_int"] == y]
     
        #     most_people_of_a_language(df_bookingtable_year, "dk")

    elif what_to_do == "dekenanalyse":
        for y in [2022,2023]:
            
            df_bookingtable_year =  df_bookingtable[df_bookingtable["year_int"] == y]
            deken_analyse(df_bookingtable_year, y)

    elif what_to_do == "babypackanalyse" :
        for y in [2022,2023]:
           
            df_bookingtable_year =  df_bookingtable[df_bookingtable["year_int"] == y]
            babypackanalyse(df_bookingtable_year, y)

    # UTILITIES
    elif what_to_do == "Find color":
        find_fill_color()
    elif what_to_do == "Compare files":
        compare_files(df_bookingtable_period, df_maxxton)
    elif what_to_do == "Check Excel File":
        check_mutation_table(df_mutation)
    elif what_to_do == "Clear cache":
        clear_cache()
    else:
        st.error("ERROR in what_to_do")


 
if __name__ == "__main__":
    print(f"________{dt.datetime.now()}_________________________")
    main()
