import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from typing import Tuple, Optional, Dict, Any
import pandas as pd

import polars as pl
import numpy as np
from pathlib import Path

# GEBRUIK C:\Users\rcxsm\Documents\python_scripts\streamlit_scripts\not_in_menu\masterfinance_streamlit.py

def load_workbook(file_path: str) -> Optional[Workbook]:
    """
    Load an Excel workbook from the given file path.

    Args:
        file_path (str): The path to the Excel file.

    Returns:
        Optional[Workbook]: The loaded workbook, or None if the file was not found.
    """
    print (f"Opening {file_path}")
    try:
        return openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: The file {file_path} was not found.")
        return None

def save_backup(workbook: Workbook, backup_path: str) -> None:
    """
    Save a backup of the given workbook to the specified path.

    Args:
        workbook (Workbook): The workbook to save.
        backup_path (str): The path to save the backup file.
    """
    print (f"Saving backup to {backup_path}")
    try:
        workbook.save(backup_path)
    except Exception as e:
        print(f"Error saving backup: {e}")

def get_sheets(workbook: Workbook) -> Tuple[Optional[Worksheet], Optional[Worksheet]]:
    """
    Get the 'INVOER' and 'RULES' sheets from the workbook.

    Args:
        workbook (Workbook): The workbook to get the sheets from.

    Returns:
        Tuple[Optional[Worksheet], Optional[Worksheet]]: The 'INVOER' and 'RULES' sheets, or (None, None) if not found.
    """
    print("Getting sheets 'INVOER' and 'RULES'")
    try:
        invoer_sheet = workbook['INVOER']
        rules_sheet = workbook['RULES']
        return invoer_sheet, rules_sheet
    except KeyError as e:
        print(f"Error: Sheet {e} not found in the workbook.")
        return None, None

def read_rules(rules_sheet: Worksheet) -> Dict[str, Tuple[Any, Any, Any, Any]]:
    """
    Read rules from the 'RULES' sheet.

    Args:
        rules_sheet (Worksheet): The 'RULES' sheet to read from.

    Returns:
        Dict[str, Tuple[Any, Any, Any, Any]]: A dictionary with rule descriptions as keys and rule details as values.
    """
    print("Reading rules from 'RULES' sheet")
    return {row[1].lower(): row[2:6] for row in rules_sheet.iter_rows(min_row=2, values_only=True) if row[1]}

def process_invoer_sheet(invoer_sheet: Worksheet, rules: Dict[str, Tuple[Any, Any, Any, Any]]) -> None:
    """
    Process the 'INVOER' sheet using the provided rules.

    Args:
        invoer_sheet (Worksheet): The 'INVOER' sheet to process.
        rules (Dict[str, Tuple[Any, Any, Any, Any]]): The rules to apply to the 'INVOER' sheet.
    """
    print ("Processing 'INVOER' sheet with rules")
    for row in invoer_sheet.iter_rows(min_row=2, min_col=8, max_col=20):
        
        description_invoer = row[4].value
        amount_invoer = row[0].value

        if description_invoer:
            try:
                description_invoer_lower = description_invoer.lower()
            except:
                description_invoer_lower = "_"
            for term, values in rules.items():
                amount_rules = values[3]

                # Check if term is in description and amount matches (if specified)
                if term in description_invoer_lower and (not amount_rules or amount_invoer == amount_rules):
                    #print(f"MATCH {description_invoer}")

                    # Only apply rules if target cells are empty
                    if all(cell.value is None for cell in (row[5], row[6], row[7])):
                        row[5].value = values[0]  # income_expenses
                        row[6].value = values[1]  # main_category
                        row[7].value = values[2]  # category
                        row[12].value = f"generated {datetime.now().strftime('%Y-%m-%d')}"
                        print(f"CHANGED MATCH {description_invoer}")
                    break  # Stop after the first match


def main_() -> None:
    """
    Main function to load the workbook, save a backup, and process the sheets.
    """
    file_path = 'C:\\Users\\rcxsm\\Documents\\xls\\masterfinance_2023.xlsx'
    backup_path = f'C:\\Users\\rcxsm\\Documents\\xls\\masterfinance_2023_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    time_start = datetime.now().strftime("%H:%M:%S")
    print("Starting workbook load at " + time_start)

  
    workbook = load_workbook(file_path)
    if not workbook:
        print("Failed to load workbook. Exiting.")
        return
    time_end = datetime.now().strftime("%H:%M:%S")
    print('Load finished at ' + time_end)
  
    save_backup(workbook, backup_path)

    invoer_sheet, rules_sheet = get_sheets(workbook)
    if not invoer_sheet or not rules_sheet:
        print("Required sheets not found. Exiting.")
        return

    rules = read_rules(rules_sheet)
    process_invoer_sheet(invoer_sheet, rules)

    workbook.save(f'C:\\Users\\rcxsm\\Documents\\xls\\masterfinance_2023_updated_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    workbook.save('C:\\Users\\rcxsm\\Documents\\xls\\masterfinance_2023.xlsx')

    time_end = datetime.now().strftime("%H:%M:%S")
    print(f"Processing completed at {time_end}.")
    
    print(f"Total time taken: {datetime.strptime(time_end, '%H:%M:%S') - datetime.strptime(time_start, '%H:%M:%S')}")

def find_kruis_posten():
    """
    Placeholder function for finding 'kruis posten'.
    Currently does nothing but can be implemented as needed.
    """
    print("Finding 'kruis posten'... ")
   
    # Lees
    path = r"C:\Users\rcxsm\Documents\xls\masterfinance_2023.xlsx"
    sheet = "INVOER"
    df = pd.read_excel(path, sheet_name=sheet)
        
    for jaar in [2020,2021,2022,2023,2024,2025]:
    #for jaar in [2025]:
        print (f"Processing {jaar}")
        out_sheet = f"KRUIS_UNMATCHED_{jaar}"
        
        # Filter
        m = (df["year"] == jaar) & (df["category"].astype(str).str.upper() == "KRUIS")
        kruis = df.loc[m].copy()

        # Normaliseer bedrag op 2 decimals
        kruis["bedrag_norm"] = kruis["Bedrag"].round(2)
        kruis["abs_key"] = kruis["bedrag_norm"].abs()

        # Markeer matches per absolute bedrag
        kruis["matched"] = False

        for key, sub in kruis.groupby("abs_key"):
            pos_idx = sub.index[sub["bedrag_norm"] > 0].tolist()
            neg_idx = sub.index[sub["bedrag_norm"] < 0].tolist()
            k = min(len(pos_idx), len(neg_idx))
            # eerste k uit beide lijsten vormen een paar
            kruis.loc[pos_idx[:k], "matched"] = True
            kruis.loc[neg_idx[:k], "matched"] = True

        unmatched = kruis.loc[~kruis["matched"]].drop(columns=["bedrag_norm","abs_key","matched"])

        # Toon resultaat
        print(f"Aantal kruisposten zonder tegenhanger: {len(unmatched)}")
        if unmatched["Bedrag"].sum() != 0:
            print(unmatched[["date","Bedrag","Description","main_category","category"]].sort_values("date"))
        
        print (f"TOTAAL BEDRAG {jaar} : {unmatched["Bedrag"].sum()}")
        if len(unmatched)>0:
            if unmatched["Bedrag"].sum() != 0:
                # Schrijf naar nieuw tabblad
                with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                    unmatched.to_excel(xw, sheet_name=out_sheet, index=False)
        else:
            print(f"Nothing found in {jaar}")

def arrow_sanitize(pdf: pd.DataFrame) -> pd.DataFrame:
    pdf = pdf.copy()
    # 1) maak alle object- en string-kolommen zuiver string
    for c in pdf.columns:
        if pd.api.types.is_object_dtype(pdf[c]) or pd.api.types.is_string_dtype(pdf[c]):
            m = ~pd.isna(pdf[c])
            pdf.loc[m, c] = pdf.loc[m, c].astype(str)
            # optioneel: forceer pandas StringDtype
            pdf[c] = pdf[c].astype("string[python]")
    return pdf

def find_kruis_posten_polars():
    path  = r"C:\Users\rcxsm\Documents\xls\masterfinance_2023.xlsx"
    sheet = "INVOER"

    print("Finding 'kruis posten'...")

    # Read via pandas then convert
    pdf = pd.read_excel(path, sheet_name=sheet)
    pdf["pos"] = None
    pdf["neg"] = None
    pdf["sign"] = None
    
    pdf = arrow_sanitize(pdf)
    
    # Repareer 'year' die soms 'YYYY-MM' is
    if "year" in pdf.columns:
        pdf["year"] = (
            pdf["year"].astype(str).str.extract(r"(\d{4})", expand=False).astype("Int64")
        )

    columns_ = [ "W",
                "source",
                "monoth",
                "year",
                "year-month",
                "saldo calculated",
                "date",
                # "Bedrag",
                "saldo",
                "bedrag_rounded",
                "bedrag_abs",
                "Description",
                "income_expenses",
                "main_category",
                "category",
                "category_asia",
                "BEDR_VR_VAL",
                "VAL",
                "KOERS",
                "SALDO VR VALUTA",
                "leeg_5",
                "leeg_6",
                "leeg_7",
                "leeg_8",
                "leeg_9",
                "leeg_10",
                "leeg_11",
                "leeg_12",
                "leeg_13",
                "leeg_14"]

    # Als je een 'Period' kolom hebt zoals '2025-04'
    for col in columns_:
        #if pdf[col].dtype.kind in "iu" and pdf[col].astype(str).str.contains(r"\d{4}-\d{2}").any():
        pdf[col] = pdf[col].astype(str)  # hou het als tekst

    df  = pl.from_pandas(pdf)

    df = df.with_columns([
        pl.col("category").cast(pl.Utf8).str.to_uppercase().alias("category_upper"),
    ])

    kruis = (
        df.filter(pl.col("category_upper") == "KRUIS")
          .with_columns([
              pl.col("Bedrag").round(2).alias("bedrag_norm"),
              pl.col("Bedrag").round(2).abs().alias("abs_key"),
              pl.when(pl.col("Bedrag").round(2) > 0).then("pos").otherwise("neg").alias("sign"),
          ])
    )

    # Optional deterministic order
    if "date" in kruis.columns:
        kruis = kruis.sort(["year","abs_key","sign","date"])

    # Rank within each group
    kruis = kruis.with_columns(
        pl.cum_count().over(["year","abs_key","sign"]).alias("rank")
    )

    # Counts and min
    cnt = (
        kruis
        .group_by(["year","abs_key","sign"])
        .len()
        .pivot(values="len", index=["year","abs_key"], columns="sign")
        .fill_null(0)
        .with_columns(
            pl.min_horizontal(["neg","pos"]).alias("min_cnt")
        )
        .select(["year","abs_key","min_cnt"])
    )

    # Join and filter unmatched
    kruis2 = (
        kruis.join(cnt, on=["year","abs_key"], how="left")
             .with_columns((pl.col("rank") < pl.col("min_cnt")).alias("matched"))
    )

    unmatched = kruis2.filter(~pl.col("matched"))
    
    # Back to pandas for Excel
    pdf_unmatched = unmatched.drop(
        ["category_upper","bedrag_norm","abs_key","sign","rank","min_cnt","matched"],
        strict=False
    ).to_pandas(use_pyarrow_extension_array=False)

    years = [2020,2021,2022,2023,2024,2025]
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
        for jaar in years:
            out_sheet = f"KRUIS_UNMATCHED_{jaar}"
            out_df = pdf_unmatched.loc[pdf_unmatched["year"] == jaar]
            print(f"Processing {jaar}  unmatched={len(out_df)}")
            if not out_df.empty:
                out_df.to_excel(xw, sheet_name=out_sheet, index=False)
            else:
                pd.DataFrame({"info":[f"Nothing found in {jaar}"]}).to_excel(
                    xw, sheet_name=out_sheet, index=False
                )


def edit_paypal():
    """ CHANGE DATE FROM MM/DD/YYYY TO DD/MM/YYYY
    """
    
    # Invoer- en uitvoerpad
    in_path = r"C:\Users\rcxsm\Downloads\A5TH6A3QBU97C-CSR-20221011000000-20251010235959-20251011100432.CSV"
    out_path = Path(in_path).with_name(Path(in_path).stem + "_dd-mm-yyyy.csv")

    # Lees alles als tekst zodat nummers met komma’s intact blijven
    df = pd.read_csv(in_path, dtype=str, encoding="utf-8-sig")

    # Converteer Datum van mm/dd/yyyy naar dd/mm/yyyy
    parsed = pd.to_datetime(df["Datum"], format="%m/%d/%Y", errors="coerce")
    df["Datum"] = parsed.dt.strftime("%d/%m/%Y").where(~parsed.isna(), df["Datum"])

    # Schrijf weg
    df.to_csv(out_path, index=False, encoding="utf-8-sig")

    print(f"Bestand opgeslagen als: {out_path}")

def main():
    time_start = datetime.now().strftime("%H:%M:%S")
    print("Starting workbook load at " + time_start)

    #main_()
    find_kruis_posten()
    #edit_paypal()
    #find_kruis_posten_polars()

    time_end = datetime.now().strftime("%H:%M:%S")
    print(f"Processing completed at {time_end}.")
    
    print(f"Total time taken: {datetime.strptime(time_end, '%H:%M:%S') - datetime.strptime(time_start, '%H:%M:%S')}")

if __name__ == "__main__":
    main()